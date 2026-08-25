"""Seed a depot's purchase history from the client's own PO and GRN books.

WHY
---
`--mode odoo-rhythm` derives supplier cadence and lead time from Odoo's own
goods receipts. It had never run against real data: the depot carried two
completed receipts in total, so the derivation was safe (it refuses to
overwrite richer data) but entirely unproven.

The client's books hold a full year of real purchase history for one store —
Rhapta Road. Loading it into the matching warehouse makes the derivation
testable end to end: derive through Odoo, then compare against the books
independently, rather than checking a file against the source it came from.

WHAT THIS CREATES, AND WHAT IT DELIBERATELY DOES NOT
----------------------------------------------------
Headers only: a purchase order with a real order date, and a receipt with a
real completion date, linked so lead time can be measured. That is exactly
what cadence and lead time are derived from.

It does NOT move stock. Every line uses a CONSUMABLE fixture product, so no
quant is written and on-hand — which the ordering engine reads — is untouched.
Seeding real quantities would rewrite the depot's stock position and invalidate
every queue built on it.

It seeds ONE warehouse, the one matching the store the books came from. The
books cover a single store; spreading them across fourteen would fabricate the
per-store variation that the cross-store discriminator is meant to measure, and
then "discover" it.

    python connectors/odoo/seed_rhythm_from_books.py --dry-run
    python connectors/odoo/seed_rhythm_from_books.py
    python connectors/odoo/seed_rhythm_from_books.py --clear
"""

from __future__ import annotations

import argparse
import glob
import os
import sys
from collections import defaultdict
from datetime import datetime

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, REPO)

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(REPO, ".env"))
except ImportError:
    pass

from oasis.logic.odoo_adapter import OdooAdapter        # noqa: E402

DATA = os.path.join(REPO, "oasis", "data")

#: Everything this script creates carries it, so a seed is findable and
#: reversible. Without a tag, cleaning up means guessing.
SEED_TAG = "OASIS-BOOKS-SEED"

#: Receipts of a CONSUMABLE create no quants. This is the whole reason the
#: seed can be loaded into a working depot without corrupting its stock.
FIXTURE_CODE = "OASIS-RHYTHM-FIXTURE"

BATCH = 200


def _hdr(ws):
    h = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(x).strip(): i for i, x in enumerate(h) if x}


def _vendor(v):
    s = str(v or "").strip()
    return (s.split(" - ", 1)[1] if " - " in s else s).strip().upper()


def _date(v):
    if isinstance(v, datetime):
        return v.date()
    for f in ("%d-%b-%Y", "%d %b %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), f).date()
        except ValueError:
            pass
    return None


def read_books(log=print):
    """(vendor, po_date, grn_date) pairs, plus receipts with no PO link."""
    import openpyxl

    po_grn = defaultdict(dict)          # vendor -> po_date -> grn_no
    for f in sorted(glob.glob(os.path.join(DATA, "po_*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        ix = _hdr(ws)
        cv, cd, cg = (ix.get("Vendor Code / Name"), ix.get("PO Date"), ix.get("GRN NO"))
        for row in ws.iter_rows(min_row=2, values_only=True):
            if cv is None:
                break
            v, d = _vendor(row[cv]), _date(row[cd])
            if v and d:
                po_grn[v][d] = str(row[cg] or "").strip() if cg is not None else ""
        wb.close()

    grn_dates = defaultdict(set)
    grn_no_date = {}
    for f in sorted(glob.glob(os.path.join(DATA, "grnds_*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        ix = _hdr(ws)
        cv, cd, cn = (ix.get("Vendor Code - Name"), ix.get("GRN Date"), ix.get("GRN No"))
        for row in ws.iter_rows(min_row=2, values_only=True):
            if cv is None:
                break
            v, d = _vendor(row[cv]), _date(row[cd])
            if not (v and d):
                continue
            grn_dates[v].add(d)
            g = str(row[cn] or "").strip() if cn is not None else ""
            if g:
                grn_no_date[g] = d
        wb.close()

    linked, loose = [], []
    claimed = defaultdict(set)
    for v, m in po_grn.items():
        for pod, gno in sorted(m.items()):
            gd = grn_no_date.get(gno) if gno else None
            if gd and 0 <= (gd - pod).days <= 180:
                linked.append((v, pod, gd))
                claimed[v].add(gd)
    for v, ds in grn_dates.items():
        for d in sorted(ds):
            if d not in claimed[v]:
                loose.append((v, d))

    log(f"books: {len(linked):,} PO->receipt pairs, {len(loose):,} receipts "
        f"with no matched PO, across "
        f"{len(set([v for v, _, _ in linked] + [v for v, _ in loose])):,} vendors")
    return linked, loose


def target_warehouse(a, log=print):
    """The warehouse matching the store the books came from."""
    whs = a._ex("stock.warehouse", "search_read", [[]],
                {"fields": ["code", "name", "in_type_id", "company_id"]}) or []
    for w in whs:
        if "RHAPTA" in str(w.get("name", "")).upper():
            log(f"target warehouse: {w['code']} — {w['name']}")
            return w
    raise SystemExit(
        "No warehouse matching the books' store (Rhapta) exists in this Odoo. "
        "Seeding into an unrelated branch would attribute one store's buying "
        "history to another; refusing.")


def fixture_product(a):
    found = a._ex("product.product", "search",
                  [[["default_code", "=", FIXTURE_CODE]]], {"limit": 1})
    if found:
        return found[0]
    return a._ex("product.product", "create", [{
        "name": "OASIS rhythm fixture (seeded history, no stock effect)",
        "default_code": FIXTURE_CODE,
        "type": "consu",
        "purchase_ok": True,
        "list_price": 0.0,
        "standard_price": 0.0,
    }])


def resolve_partners(a, names, log=print):
    have = {}
    for i in range(0, len(names), 500):
        chunk = names[i:i + 500]
        for p in (a._ex("res.partner", "search_read", [[["name", "in", chunk]]],
                        {"fields": ["name"]}) or []):
            have[str(p["name"]).strip().upper()] = p["id"]
    missing = [n for n in names if n not in have]
    log(f"partners: {len(have):,} matched, {len(missing):,} to create")
    for i in range(0, len(missing), BATCH):
        chunk = missing[i:i + BATCH]
        ids = a._ex("res.partner", "create",
                    [[{"name": n, "supplier_rank": 1, "company_type": "company"}
                      for n in chunk]])
        for n, pid in zip(chunk, ids if isinstance(ids, list) else [ids]):
            have[n] = pid
    return have


def clear(a, log=print):
    picks = a._ex("stock.picking", "search", [[["origin", "like", SEED_TAG]]])
    if picks:
        for i in range(0, len(picks), BATCH):
            a._ex("stock.picking", "write",
                  [picks[i:i + BATCH], {"state": "draft"}])
            a._ex("stock.picking", "unlink", [picks[i:i + BATCH]])
    pos = a._ex("purchase.order", "search", [[["origin", "like", SEED_TAG]]])
    if pos:
        for i in range(0, len(pos), BATCH):
            # CANCEL, not draft. Odoo's _unlink_if_cancelled refuses to delete a
            # purchase order in any other state — the same rule the ordering
            # module's lifecycle tests pin down.
            a._ex("purchase.order", "write",
                  [pos[i:i + BATCH], {"state": "cancel"}])
            a._ex("purchase.order", "unlink", [pos[i:i + BATCH]])
    log(f"removed {len(picks):,} seeded receipts and {len(pos):,} seeded orders")


def seed(a, linked, loose, wh, prod, partners, log=print):
    ptype = wh["in_type_id"][0]
    company = wh["company_id"][0]
    made_po = made_pick = 0

    # ── linked: a purchase order and the receipt that closed it ───────────
    for i in range(0, len(linked), BATCH):
        chunk = linked[i:i + BATCH]
        po_vals = []
        for v, pod, gd in chunk:
            po_vals.append({
                "partner_id": partners[v],
                "company_id": company,
                "picking_type_id": ptype,
                "date_order": f"{pod} 09:00:00",
                "origin": SEED_TAG,
                "order_line": [(0, 0, {
                    "product_id": prod, "name": "seeded history",
                    "product_qty": 1, "price_unit": 0.0,
                    "date_planned": f"{pod} 09:00:00"})],
            })
        po_ids = a._ex("purchase.order", "create", [po_vals])
        po_ids = po_ids if isinstance(po_ids, list) else [po_ids]
        made_po += len(po_ids)

        lines = {r["order_id"][0]: r["id"] for r in
                 (a._ex("purchase.order.line", "search_read",
                        [[["order_id", "in", po_ids]]], {"fields": ["order_id"]}) or [])}

        pick_vals = [{"partner_id": partners[v], "picking_type_id": ptype,
                      "company_id": company,
                      "scheduled_date": f"{gd} 09:00:00",
                      "origin": f"{SEED_TAG} {po_id}"}
                     for (v, pod, gd), po_id in zip(chunk, po_ids)]
        pick_ids = a._ex("stock.picking", "create", [pick_vals])
        pick_ids = pick_ids if isinstance(pick_ids, list) else [pick_ids]

        move_vals = []
        for (v, pod, gd), po_id, pk in zip(chunk, po_ids, pick_ids):
            if po_id not in lines:
                continue
            move_vals.append({
                "name": "seeded history", "product_id": prod,
                "product_uom_qty": 1, "picking_id": pk,
                "purchase_line_id": lines[po_id],
                "location_id": 4, "location_dest_id": 8,
                "company_id": company,
            })
        if move_vals:
            a._ex("stock.move", "create", [move_vals])

        # One write per DISTINCT date, not per record. Receipts cluster heavily
        # on the same days, so this turns ~9,000 round trips into a few hundred.
        by_day = defaultdict(list)
        for (v, pod, gd), pk in zip(chunk, pick_ids):
            by_day[gd].append(pk)
        for gd, pks in by_day.items():
            a._ex("stock.picking", "write",
                  [pks, {"state": "done", "date_done": f"{gd} 09:00:00"}])
        made_pick += len(pick_ids)
        log(f"   linked {min(i + BATCH, len(linked)):,}/{len(linked):,}")

    # ── loose receipts: cadence only, no lead time to measure ─────────────
    for i in range(0, len(loose), BATCH):
        chunk = loose[i:i + BATCH]
        vals = [{"partner_id": partners[v], "picking_type_id": ptype,
                 "company_id": company, "scheduled_date": f"{d} 09:00:00",
                 "origin": SEED_TAG} for v, d in chunk]
        ids = a._ex("stock.picking", "create", [vals])
        ids = ids if isinstance(ids, list) else [ids]
        by_day = defaultdict(list)
        for (v, d), pk in zip(chunk, ids):
            by_day[d].append(pk)
        for d, pks in by_day.items():
            a._ex("stock.picking", "write",
                  [pks, {"state": "done", "date_done": f"{d} 09:00:00"}])
        made_pick += len(ids)
        log(f"   loose {min(i + BATCH, len(loose)):,}/{len(loose):,}")

    return made_po, made_pick


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--clear", action="store_true")
    p.add_argument("--limit", type=int, default=0,
                   help="cap events (for a quick rehearsal; distorts cadence)")
    args = p.parse_args(argv)

    a = OdooAdapter()
    if not a.health_check().get("connected"):
        raise SystemExit("Odoo unreachable — is the stack up?")

    if args.clear:
        clear(a)
        return 0

    linked, loose = read_books()
    if args.limit:
        linked, loose = linked[:args.limit], loose[:args.limit]

    wh = target_warehouse(a)
    before = a._ex("stock.quant", "search_count", [[]])

    if args.dry_run:
        print("\n--dry-run — nothing written.")
        print(f"  would create {len(linked):,} purchase orders")
        print(f"  would create {len(linked) + len(loose):,} completed receipts")
        print(f"  into {wh['code']}, all tagged {SEED_TAG!r}")
        print(f"  stock quants would stay at {before:,} (consumable fixture)")
        return 0

    names = sorted({v for v, _, _ in linked} | {v for v, _ in loose})
    partners = resolve_partners(a, names)
    prod = fixture_product(a)
    made_po, made_pick = seed(a, linked, loose, wh, prod, partners)

    after = a._ex("stock.quant", "search_count", [[]])
    print(f"\n-> {made_po:,} purchase orders, {made_pick:,} completed receipts "
          f"into {wh['code']}")
    print(f"   stock quants {before:,} -> {after:,} "
          f"({'untouched' if before == after else 'CHANGED — investigate'})")
    print("   now: python entrypoint.py --mode odoo-rhythm")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
