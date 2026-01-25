
import os
import glob
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple
from openpyxl import load_workbook
from statistics import median, mean
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "app", "data")
if not os.path.exists(DATA_DIR):
    DATA_DIR = os.path.join(os.getcwd(), "app", "data")

def parse_date(date_val: Any) -> datetime:
    """Parse date from various formats."""
    if isinstance(date_val, datetime):
        return date_val
    if not date_val:
        return None
    
    date_str = str(date_val).strip()
    formats = [
        '%d-%b-%Y', '%d-%m-%Y', '%Y-%m-%d', 
        '%d-%b-%y', '%m/%d/%Y', '%d/%m/%Y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

class CalendarAnalyzer:
    def __init__(self):
        self.po_data = []  # List of (Supplier, PO Number, PO Date)
        self.grn_data = {} # Dict of PO Number -> GRN Date
        self.suppliers = {} # Dict of Supplier -> Metadata

    def load_data(self):
        """Load PO and GRN data from Excel files."""
        self._load_po_files()
        self._load_grn_files()
        logger.info(f"Loaded {len(self.po_data)} POs and {len(self.grn_data)} GRN records.")

    def _load_po_files(self):
        po_files = glob.glob(os.path.join(DATA_DIR, "po_*.xlsx"))
        logger.info(f"Found {len(po_files)} PO files.")
        
        for fpath in po_files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                
                # Headers are in row 1, data starts row 2
                # Columns: C=Supplier(2), D=Date(3), E=PO No(4)
                # 0-indexed: Supplier=2, Date=3, PO No=4
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    supplier = row[2]
                    po_date_raw = row[3]
                    po_no = str(row[4]).strip() if row[4] else None
                    
                    if supplier and po_date_raw and po_no:
                        po_date = parse_date(po_date_raw)
                        if po_date:
                            self.po_data.append({
                                'supplier': str(supplier).strip().upper(),
                                'po_no': po_no,
                                'date': po_date
                            })
            except Exception as e:
                logger.error(f"Error loading PO file {os.path.basename(fpath)}: {e}")

    def _load_grn_files(self):
        grn_files = glob.glob(os.path.join(DATA_DIR, "grnds_*.xlsx"))
        logger.info(f"Found {len(grn_files)} GRN files.")
        
        for fpath in grn_files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                
                # GRN files headers usually in row 1
                # Check check_grn_leadtime_headers.py output:
                # GRN Date is col 2 (0-indexed), PO No is col 4
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    grn_date_raw = row[2]
                    po_no = str(row[4]).strip() if row[4] else None
                    
                    if grn_date_raw and po_no:
                        grn_date = parse_date(grn_date_raw)
                        if grn_date:
                            # Verify if we already have this PO (sometimes partial deliveries)
                            # We'll take the LAST GRN date for lead time conservatism
                            if po_no not in self.grn_data or grn_date > self.grn_data[po_no]:
                                self.grn_data[po_no] = grn_date
            except Exception as e:
                logger.error(f"Error loading GRN file {os.path.basename(fpath)}: {e}")

    def analyze(self) -> Dict[str, Any]:
        """Analyze data to determine frequency and lead times."""
        grouped = {}
        
        # Group POs by supplier
        for entry in self.po_data:
            s_name = entry['supplier']
            if s_name not in grouped:
                grouped[s_name] = {'dates': [], 'lead_times': []}
            
            grouped[s_name]['dates'].append(entry['date'])
            
            # Calculate lead time if GRN exists
            if entry['po_no'] in self.grn_data:
                grn_date = self.grn_data[entry['po_no']]
                delta = (grn_date - entry['date']).days
                if 0 <= delta <= 60: # Sanity check
                    grouped[s_name]['lead_times'].append(delta)

        results = {}
        
        for supplier, data in grouped.items():
            dates = sorted(list(set(data['dates'])))
            lead_times = data['lead_times']
            
            if len(dates) < 2:
                freq_val = 30 # Default to monthly if not enough data
            else:
                intervals = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
                freq_val = median(intervals) if intervals else 30
            
            avg_lead_time = int(mean(lead_times)) if lead_times else 3 # Default 3 days
            
            # Classification
            if freq_val < 3.5:
                category = 'Daily'
                preferred_day = None
            elif freq_val < 11:
                category = 'Weekly'
                # Find preferred weekday (0=Mon, 6=Sun)
                weekdays = [d.weekday() for d in dates]
                preferred_day = max(set(weekdays), key=weekdays.count)
            elif freq_val < 25:
                category = 'Bi-Weekly'
                weekdays = [d.weekday() for d in dates]
                preferred_day = max(set(weekdays), key=weekdays.count)
            else:
                category = 'Monthly'
                # Find preferred day of month
                doms = [d.day for d in dates]
                preferred_day = max(set(doms), key=doms.count)
            
            results[supplier] = {
                'frequency_days': freq_val,
                'lead_time_days': avg_lead_time,
                'category': category,
                'preferred_day': preferred_day, # 0-6 for weekly, 1-31 for monthly
                'order_count': len(dates)
            }
            
        return results

if __name__ == "__main__":
    analyzer = CalendarAnalyzer()
    analyzer.load_data()
    res = analyzer.analyze()
    for s, data in list(res.items())[:5]:
        print(f"{s}: {data}")
