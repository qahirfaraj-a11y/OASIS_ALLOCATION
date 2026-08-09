import os
import csv
import logging
from datetime import datetime
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook

import warnings
# Suppress openpyxl warnings (v10.0 Optimization)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

logger = logging.getLogger("OrderEngine.Data")

class DataMixin:
    """
    DataMixin handles all file I/O operations including parsing inventory files (Excel/CSV)
    and generating Excel reports.
    """
    # Type hints for attributes provided by the base OrderEngine
    data_dir: str
    databases: Dict[str, Any]
    grn_db: Dict[str, Any]
    grn_frequency_map: Dict[str, float]
    
    def _safe_int(self, value: Any) -> int:
        try:
            if value is None: return 0
            return int(float(str(value).strip().replace(',', '')))
        except:
            return 0

    def _safe_float(self, value: Any) -> float:
        try:
            if value is None: return 0.0
            return float(str(value).strip().replace(',', ''))
        except:
            return 0.0

    # load_databases_async removed: Handled by OrderEngine.load_databases_async

    def parse_inventory_file(self, file_path: str) -> List[dict]:
        """Phase 1: File Parsing. Supports CSV, Excel, and Picking list formats."""
        logger.info(f"Phase 1: Parsing inventory file: {file_path}")
        
        products = []
        is_excel = file_path.lower().endswith(('.xlsx', '.xls'))
        
        try:
            if is_excel:
                # Load with read_only=True for memory efficiency (v2026 Optimization)
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                
                # 1. Extract supplier from Row 1, Col 7-10
                supplier_name = None
                for col in range(7, 11):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and not supplier_name:
                        supplier_name = str(cell_value).strip()
                        break
                if not supplier_name:
                    supplier_name = 'UNKNOWN SUPPLIER'
                
                # 2. Extract headers from Row 3
                headers = [ws.cell(row=3, column=col).value for col in range(1, 30)]
                headers = [str(h).strip().upper() if h is not None else '' for h in headers]
                
                # Map columns (strictly as per picking list specs)
                col_map = {h: i+1 for i, h in enumerate(headers) if h}
                
                # Check if it's a standard format or Picking List
                # v2025: Strictly check for RR PREV and ANCHOR to trigger authoritative logic
                is_picking_list = ('RR PREV' in col_map) and ('ANCHOR' in col_map)
                
                if is_picking_list:
                    # 3. Parse product data starting Row 4
                    for row_idx in range(4, ws.max_row + 1):
                        p_name = ws.cell(row=row_idx, column=col_map.get('DESCRIPTION', 1)).value
                        if not p_name: continue

                        rr_prev = self._safe_float(ws.cell(row=row_idx, column=col_map.get('RR PREV', 0)).value) if col_map.get('RR PREV') else 0.0
                        pb_val = str(ws.cell(row=row_idx, column=col_map.get('RR PB', 0)).value).strip().upper() if col_map.get('RR PB') else '0'
                        blocked_status = 'blocked' if pb_val in ['1', 'BLOCKED', '1.0'] else 'open'
                        
                        product = {
                            "product_name": str(p_name).strip(),
                            "item_code": str(ws.cell(row=row_idx, column=col_map.get('ITEM CODE', 0)).value or '').strip() if col_map.get('ITEM CODE') else '',
                            "barcode": str(ws.cell(row=row_idx, column=col_map.get('BARCODE', 0)).value or '').strip() if col_map.get('BARCODE') else '',
                            "supplier_name": supplier_name,
                            "current_stocks": self._safe_float(ws.cell(row=row_idx, column=col_map.get('ANCHOR', 0)).value) if col_map.get('ANCHOR') else 0.0,
                            "units_sold_last_month": rr_prev,
                            "estimated_daily_sales": rr_prev / 30.0 if rr_prev > 0 else 0.0,
                            "last_days_since_last_delivery": self._safe_int(ws.cell(row=row_idx, column=col_map.get('RR GRN', 0)).value) if col_map.get('RR GRN') else 0,
                            "blocked_open_for_order": blocked_status,
                            "pack_size": self._safe_int(ws.cell(row=row_idx, column=col_map.get('PACK', 0)).value) if col_map.get('PACK') else 1,
                            "selling_price": float(self._safe_float(ws.cell(row=row_idx, column=col_map.get('SP', 0)).value)) if col_map.get('SP') else 0.0,
                            "product_category": 'general',
                            "is_fresh": any(k in str(p_name).upper() for k in ['MILK', 'BREAD', 'DAIRY', 'YOGURT', 'CAKE', 'ROLL']),
                            # Preserving original picking list field names for report generator
                            "RR PREV": rr_prev,
                            "RR GRN": self._safe_int(ws.cell(row=row_idx, column=col_map.get('RR GRN', 0)).value) if col_map.get('RR GRN') else 0,
                            "RR PB": pb_val,
                            # Client spreadsheet column header — theirs, not ours. Renaming it
                            # to de-identify would simply stop us reading their file.
                            "RHAPTA": self._safe_float(ws.cell(row=row_idx, column=col_map.get('RHAPTA', 0)).value) if col_map.get('RHAPTA') else 0.0,
                            "Space": ws.cell(row=row_idx, column=col_map.get('SPACE', 0)).value if col_map.get('SPACE') else ''
                        }
                        products.append(product)
                else:
                    # Standard Excel logic (headers in row 1)
                    headers = [str(cell.value).strip().lower().replace(' ', '_') for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            products.append(dict(zip(headers, row)))
            else:
                # CSV parsing with encoding failover (v10.0 Robustness)
                for enc in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        with open(file_path, 'r', encoding=enc) as f:
                            reader = csv.DictReader(f)
                            products = list(reader)
                        logger.info(f"CSV parsed successfully with {enc}")
                        break
                    except Exception as e:
                        if enc == 'cp1252': raise e
                        continue

        except Exception as e:
            logger.error(f"Error parsing file: {e}")
            raise
            
        # Standardize field names (v10.1 robustness against Csv Casing)
        for p in products:
            # Create a lower-case alias map for safe fetching
            p_lower = {str(k).lower(): v for k, v in p.items()}
            
            p['product_name'] = (p.get('product_name') or p.get('description') or 
                                 p_lower.get('item_name') or p_lower.get('itm_name') or 
                                 p_lower.get('product') or 'Unknown')
            
            p['current_stocks'] = self._safe_float(
                p.get('current_stocks') or p_lower.get('current_stock') or p_lower.get('stock_on_hand') or 
                p_lower.get('soh') or p_lower.get('anchor_store_soh')
                or p_lower.get('027_-_rhapta_road') or 0
            )
            
            # v10.10: Capture Live Sales Data from Forensic Exports
            live_units_30d = self._safe_float(p_lower.get('total_units_sold_(30_days)') or p_lower.get('units_sold_30d') or 0)
            live_ads = live_units_30d / 30.0 if live_units_30d > 0 else 0.0

            p['avg_daily_sales'] = self._safe_float(
                p.get('avg_daily_sales') or p_lower.get('avg_daily_sales') or p_lower.get('daily_usage') or 
                p_lower.get('ads') or p_lower.get('estimated_daily_sales') or live_ads or 0
            )
            
            # Store raw live data for blending in IntelligenceMixin
            p['live_ads_30d'] = live_ads

            p['cost_price'] = self._safe_float(
                p.get('cost_price') or p_lower.get('unit_cost') or p_lower.get('cost') or p_lower.get('unit_price') or 0
            )

            p['item_code'] = p.get('item_code') or p_lower.get('code') or p_lower.get('itm_code') or p_lower.get('itm_cd')
            p['barcode'] = p.get('barcode') or p_lower.get('barcode')
            p['supplier_name'] = p.get('supplier_name') or p_lower.get('vendor') or p_lower.get('supplier') or 'Unknown'
            p['department'] = p.get('department') or p_lower.get('dept') or p_lower.get('department') or 'GENERAL'
            
            p['last_days_since_last_delivery'] = self._safe_int(
                p.get('last_days_since_last_delivery') or p_lower.get('last_delivery_days') or p_lower.get('rr_grn') or 0
            )
            
            p['blocked_open_for_order'] = (p.get('blocked_open_for_order') or 
                                           p_lower.get('blocked_status', 'open') or 
                                           ('blocked' if p.get('rr_pb') in ['1', 'BLOCKED'] else 'open'))
            
        return products

    def _load_products(self, search_dir: Optional[str] = None) -> dict:
        """
        [GOLDEN LOGIC v10.0] 
        Comprehensive GRN scanner for harvesting historical cost and frequency patterns.
        Populates self.grn_db and self.grn_frequency_map.
        """
        import glob
        target_dir = search_dir or self.data_dir
        logger.info(f"Phase 3: Deep-Scanning GRNs in {target_dir}...")
        grn_stats = {}
        sku_order_dates = {}
        
        # Pattern matches both grnd_ and grnds_
        files = glob.glob(os.path.join(target_dir, "grnd*.xlsx"))
        
        for fpath in files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row: continue
                
                headers = {str(val).strip().lower().replace(' ', '').replace('_', ''): idx for idx, val in enumerate(header_row) if val}
                
                col_qty = headers.get('grnqty', headers.get('qty'))
                col_cost = headers.get('avgcost', headers.get('cost', headers.get('unitprice')))
                col_barcode = headers.get('barcode')
                col_name = headers.get('itemname', headers.get('description'))
                col_code = headers.get('itemcode', headers.get('code'))
                col_date = headers.get('grndate', headers.get('docdate', headers.get('date')))
                
                if col_qty is None: continue

                for row in ws.iter_rows(min_row=2, values_only=True):
                    qty = self._safe_float(row[col_qty])
                    if qty <= 0: continue
                    
                    cost = self._safe_float(row[col_cost]) if col_cost is not None else 0.0
                    barcode = str(row[col_barcode]).strip() if col_barcode is not None and row[col_barcode] else None
                    name = str(row[col_name]).strip().upper() if col_name is not None and row[col_name] else None
                    code = str(row[col_code]).strip() if col_code is not None and row[col_code] else None
                    
                    # Date Harvesting for Frequency
                    date_val = row[col_date] if col_date is not None else None
                    date_obj = None
                    if isinstance(date_val, datetime):
                        date_obj = date_val
                    elif isinstance(date_val, str):
                        for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%m/%d/%Y'):
                            try:
                                date_obj = datetime.strptime(date_val, fmt)
                                break
                            except: continue

                    keys = [k for k in [barcode, name, code] if k]
                    for key in keys:
                        # Update Stats
                        if key not in grn_stats: 
                            grn_stats[key] = {'total': 0.0, 'count': 0, 'avg_cost': 0.0}
                        s = grn_stats[key]
                        existing_total = s['total']
                        new_total = existing_total + qty
                        if cost > 0:
                            current_avg = s['avg_cost']
                            s['avg_cost'] = ((current_avg * existing_total) + (cost * qty)) / new_total
                        s['total'] = new_total
                        s['count'] += 1
                        
                        # Update Dates for Rhythm
                        if date_obj:
                            if key not in sku_order_dates: sku_order_dates[key] = set()
                            sku_order_dates[key].add(date_obj.date())
                wb.close()
            except Exception as e:
                logger.error(f"Error reading GRN {os.path.basename(fpath)}: {e}")
        
        # Calculate Frequency (Inverse of Gap)
        self.grn_frequency_map = {}
        for key, dates in sku_order_dates.items():
            if len(dates) > 1:
                sorted_dates = sorted(list(dates))
                gaps = [(sorted_dates[i] - sorted_dates[i-1]).days for i in range(1, len(sorted_dates))]
                avg_gap = sum(gaps) / len(gaps)
                if avg_gap > 0:
                    self.grn_frequency_map[key] = 1.0 / avg_gap
        
        # BUG 8 FIX: Populate days_since_last_grn for each SKU.
        # IntelligenceMixin reads grn_stat.get('days_since_last_grn') for discontinued detection,
        # but the scanner never set this field — all items appeared as 0 days old.
        from datetime import date as date_type
        today = date_type.today()
        for key, dates in sku_order_dates.items():
            if key in grn_stats and dates:
                max_date = max(dates)
                grn_stats[key]['days_since_last_grn'] = (today - max_date).days
        
        logger.info(f"Scanning Complete. Harvesting Stats for {len(grn_stats)} SKUs and Rhythm for {len(self.grn_frequency_map)} SKUs.")
        return grn_stats

    def scan_purchase_orders(self) -> dict:
        """
        Scans all purchase order Excel files (po_*.xlsx) and extracts supplier ordering history.
        Used to calculate order rhythm (gaps between orders).
        """
        import glob
        logger.info("Scanning Purchase Order Excel files for rhythm tracking...")
        po_history = {}
        
        files = glob.glob(os.path.join(self.data_dir, "po_*.xlsx"))
        
        for fpath in files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row: continue
                
                headers = {"".join(c for c in str(val).lower() if c.isalnum()): idx for idx, val in enumerate(header_row) if val}
                col_vendor = headers.get('vendorcodename', headers.get('vendor'))
                col_date = headers.get('podate', headers.get('date'))

                for row in ws.iter_rows(min_row=2, values_only=True):
                    vendor_raw = str(row[col_vendor]).strip() if col_vendor is not None and row[col_vendor] else None
                    if not vendor_raw: continue
                    
                    # Normalize: "SB0009 - BROOKSIDE DAIRY" -> "BROOKSIDE DAIRY"
                    supplier = vendor_raw.split(' - ', 1)[1].upper().strip() if ' - ' in vendor_raw else vendor_raw.upper().strip()
                    date_val = row[col_date]
                    if not date_val: continue
                    
                    if isinstance(date_val, str):
                        for fmt in ('%d-%b-%Y', '%Y-%m-%d', '%m/%d/%Y'):
                            try:
                                date_obj = datetime.strptime(date_val, fmt)
                                break
                            except ValueError: continue
                        else: continue
                    elif isinstance(date_val, datetime):
                        date_obj = date_val
                    else: continue

                    if supplier not in po_history: po_history[supplier] = []
                    po_history[supplier].append(date_obj)
                wb.close()
            except Exception as e:
                logger.error(f"Error reading PO {os.path.basename(fpath)}: {e}")
        
        for s in po_history: po_history[s].sort()
        return po_history

    def scan_purchase_returns(self) -> dict:
        """Scans PRTS files to aggregate supplier quality data."""
        import glob
        logger.info("Scanning Purchase Return files...")
        return_stats = {}
        files = glob.glob(os.path.join(self.data_dir, "prts_*.xlsx"))
        
        for fpath in files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                headers = {"".join(c for c in str(v).lower() if c.isalnum()): idx for idx, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))}
                col_supp = headers.get('vencodename', headers.get('vendor'))
                col_qty = headers.get('rejcqty', headers.get('qty'))
                col_amt = headers.get('netamt', headers.get('amount'))
                col_reason = headers.get('reason')

                for row in ws.iter_rows(min_row=2, values_only=True):
                    supp_raw = str(row[col_supp]) if col_supp is not None and row[col_supp] else None
                    if not supp_raw: continue
                    supplier = supp_raw.split(' - ', 1)[1].upper().strip() if ' - ' in supp_raw else supp_raw.upper().strip()
                    
                    if supplier not in return_stats:
                        return_stats[supplier] = {'total_returns': 0, 'expiry_returns': 0, 'damaged_returns': 0, 'total_value_returned': 0.0, 'total_qty_returned': 0.0, 'short_supply_returns': 0}
                    
                    stats = return_stats[supplier]
                    stats['total_returns'] += 1
                    stats['total_value_returned'] += self._safe_float(row[col_amt])
                    stats['total_qty_returned'] += self._safe_float(row[col_qty])
                    
                    reason = str(row[col_reason]).upper() if col_reason is not None else ""
                    if 'EXPIRY' in reason or 'EXP' in reason: stats['expiry_returns'] += 1
                    elif 'DAMAGE' in reason: stats['damaged_returns'] += 1
                    elif 'SHORT' in reason or 'SUPPLY' in reason:
                        if 'short_supply_returns' not in stats: stats['short_supply_returns'] = 0
                        stats['short_supply_returns'] += 1
                wb.close()
            except Exception as e: logger.error(f"Error reading PRTS {fpath}: {e}")
        return return_stats

    def scan_cashier_sales(self) -> dict:
        """Scans cashier POS sales Excel files (*_cash.xlsx)."""
        import glob
        logger.info("Scanning Cashier POS Sales...")
        sales = {}
        files = glob.glob(os.path.join(self.data_dir, "*_cash.xlsx"))
        for fpath in files:
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = ws.iter_rows(min_row=2, values_only=True)
                header = next(rows, None)
                if not header: continue
                cols = {str(v).strip().lower(): i for i, v in enumerate(header) if v}
                c_code = cols.get('itm code', cols.get('code'))
                c_qty = cols.get('qty')
                c_name = cols.get('item name', cols.get('description'))
                for row in rows:
                    qty = self._safe_float(row[c_qty])
                    if qty <= 0: continue
                    for key in [str(row[c_code]).strip() if c_code is not None else None, str(row[c_name]).strip().upper() if c_name is not None else None]:
                        if key: sales[key] = sales.get(key, 0.0) + qty
                wb.close()
            except Exception as e: logger.error(f"Error reading Sales {fpath}: {e}")
        return sales

    def scan_inventory_transfers(self) -> dict:
        """Scans transfer in/out files (trn_*.xlsx and trout_*.xlsx)."""
        import glob
        logger.info("Scanning Inventory Transfers...")
        transfers = {}
        for pattern, key_type in [("trn_*.xlsx", "in"), ("trout_*.xlsx", "out")]:
            for fpath in glob.glob(os.path.join(self.data_dir, pattern)):
                try:
                    wb = load_workbook(fpath, read_only=True, data_only=True)
                    ws = wb.active
                    headers = {str(v).strip().lower().replace(' ', ''): idx for idx, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))}
                    c_bc = headers.get('barcode')
                    c_qty = headers.get('stiqty', headers.get('stoqty', headers.get('qty')))
                    c_name = headers.get('itemname')
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        qty = self._safe_float(row[c_qty])
                        for key in [str(row[c_bc]).strip() if c_bc is not None else None, str(row[c_name]).strip().upper() if c_name is not None else None]:
                            if key:
                                if key not in transfers: transfers[key] = {'in': 0.0, 'out': 0.0}
                                transfers[key][key_type] += qty
                    wb.close()
                except Exception as e: logger.error(f"Error reading Transfer {fpath}: {e}")
        return transfers

    def load_grn_frequency(self) -> Dict[str, float]:
        """Loads historical GRN frequency tracking from JSON."""
        freq_path = os.path.join(self.data_dir, 'sku_grn_frequency.json')
        if os.path.exists(freq_path):
            try:
                import json
                with open(freq_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    logger.info(f"Loaded GRN frequencies for {len(data)} items.")
                    return {str(k).upper(): float(v) for k, v in data.items()}
            except Exception as e:
                logger.error(f"Error loading GRN frequency: {e}")
        return {}

    def generate_excel_report(self, input_file: str, recommendations: List[dict], output_path: str):
        """Phase 6: Report Generation. Standardized 15-Column Layout with Summary."""
        logger.info(f"Phase 6: Generating Excel report at {output_path}")
        
        try:
            # Non-Destructive Saving (Optimization v7.9)
            if not os.path.exists(output_path):
                import shutil
                shutil.copy2(input_file, output_path)
            
            # Load original workbook
            wb = load_workbook(output_path)
            ws = wb.active # Assuming single sheet or first sheet relevant
            
            # Map recommendations by product name for O(1) lookup
            rec_map = {r['product_name']: r for r in recommendations}
            
            # Find header row (Search for common keywords like "Item Name" or "Department")
            header_row_idx = 1
            found_header = False
            for r in range(1, 10):
                for c in range(1, 10):
                    val = str(ws.cell(row=r, column=c).value or '').strip().lower()
                    if 'item' in val or 'dept' in val or 'description' in val:
                        header_row_idx = r
                        found_header = True
                        break
                if found_header: break
            
            logger.info(f"Header row detected at index {header_row_idx}")
            
            col_map = {}
            for col in range(1, 40):
                val = ws.cell(row=header_row_idx, column=col).value
                if val:
                    col_map[str(val).strip().lower().replace(' ', '_')] = col
            
            # Add new headers
            new_headers = ["Recommended Qty", "Historical Avg", "Confidence", "Reasoning", "Est. Cost (KES)"]
            
            # Use fixed columns 11-15 for "Picking List" format if description is at Col 1
            is_picking_list = col_map.get('rr_prev') and col_map.get('description') == 1
            
            if is_picking_list:
                start_col = 11
            else:
                # Find first empty column after existing headers
                start_col = 1
                for col in range(1, 100):
                    val = ws.cell(row=header_row_idx, column=col).value
                    if not val:
                        start_col = col
                        break
                    else:
                        start_col = col + 1
            
            for i, h in enumerate(new_headers):
                c = ws.cell(row=header_row_idx, column=start_col + i)
                c.value = h
                c.font = c.font.copy(bold=True)
                if is_picking_list:
                    from openpyxl.styles import PatternFill
                    c.fill = PatternFill(start_color="4A9EFF", end_color="4A9EFF", fill_type="solid")
                    c.font = c.font.copy(color="FFFFFF")

            desc_col = col_map.get('description', col_map.get('product_name', col_map.get('item_name')))
            
            if not desc_col:
                logger.error("Could not find Description/Product Name column.")
                return

            total_rec_units = 0
            total_est_cost = 0.0

            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                product_name_cell = ws.cell(row=row_idx, column=desc_col).value
                if not product_name_cell: continue
                
                product_name = str(product_name_cell).strip()
                rec = rec_map.get(product_name, {})
                
                qty = float(rec.get('recommended_quantity', 0))
                hist = rec.get('historical_avg_order_qty', 0)
                conf = rec.get('confidence_grn', 'LOW')
                reason = rec.get('reasoning', '')
                
                # Derive est_cost or manually compute to prevent missing data
                cost_price = float(rec.get('cost_price', rec.get('selling_price', 0) * 0.75))
                cost = rec.get('est_cost', qty * cost_price)
                
                # Write values
                ws.cell(row=row_idx, column=start_col).value = qty
                ws.cell(row=row_idx, column=start_col + 1).value = hist
                ws.cell(row=row_idx, column=start_col + 2).value = conf
                ws.cell(row=row_idx, column=start_col + 3).value = reason
                ws.cell(row=row_idx, column=start_col + 4).value = cost
                
                total_rec_units += qty
                total_est_cost += cost

            # Create Summary Sheet (Phase 6)
            if "Order Summary" in wb.sheetnames:
                del wb["Order Summary"]
            ws_summary = wb.create_sheet("Order Summary", 0)
            
            est_savings = total_est_cost * 0.10 # 10% Waste Reduction
            
            summary_data = [
                ["Metric", "Value"],
                ["Total Products Analyzed", len(recommendations)],
                ["Total Recommended Units", total_rec_units],
                ["Estimated Total Cost (KES)", f"{total_est_cost:,.2f}"],
                ["Estimated Savings (10% Waste Red.)", f"{est_savings:,.2f}"],
                ["Generated On", "AI Inventory Assistant"]
            ]
            
            for r_idx, r_data in enumerate(summary_data, 1):
                ws_summary.cell(row=r_idx, column=1).value = r_data[0]
                ws_summary.cell(row=r_idx, column=2).value = r_data[1]

            # Apply Filter to new columns (Optimization v10.0)
            import openpyxl
            ws.auto_filter.ref = f"A{header_row_idx}:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"

            wb.save(output_path)
            logger.info("Excel report saved successfully.")
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            raise

    def update_supplier_quality_scores(self):
        """v10.0 Algorithm: Integrates PRTS return data into Supplier Quality Scorecard (SQS)."""
        logger.info("Updating Supplier Quality Scores...")
        return_data = self.scan_purchase_returns()
        if not return_data: return
        
        import json
        sqs_path = os.path.join(self.data_dir, 'supplier_quality_scores_2025.json')
        sqs_db = {}
        if os.path.exists(sqs_path):
            try:
                with open(sqs_path, 'r', encoding='utf-8') as f:
                    sqs_db = json.load(f)
            except Exception as e:
                logger.error(f"Failed to read SQS DB: {e}")
                
        updates_made = 0
        for supplier, stats in return_data.items():
            if supplier not in sqs_db:
                # Default baseline = 85.0
                sqs_db[supplier] = {"quality_score": 85.0, "notes": "Baseline SQS", "delivery_reliability": 0.9}
            
            # Penelope Penalties
            penalty = 0.0
            penalty += stats.get('expiry_returns', 0) * 0.5  # Heavy penalty for expiry
            penalty += stats.get('damaged_returns', 0) * 0.2
            penalty += stats.get('short_supply_returns', 0) * 0.3 # Moderate penalty for missing order lines
            
            # Deduct from baseline (max deduction of 40 points)
            current_score = float(sqs_db[supplier].get('quality_score', 85.0))
            new_score = max(45.0, current_score - penalty)
            
            # v10.0 Scale delivery_reliability proportionally
            new_reliability = max(0.4, sqs_db[supplier].get('delivery_reliability', 0.9) - (penalty * 0.01))
            
            if abs(current_score - new_score) > 0.1:
                sqs_db[supplier]['quality_score'] = round(new_score, 1)
                sqs_db[supplier]['delivery_reliability'] = round(new_reliability, 3)
                sqs_db[supplier]['notes'] = f"Adjusted by automation due to {stats.get('total_returns', 0)} returns."
                updates_made += 1
                
        if updates_made > 0:
            try:
                with open(sqs_path, 'w', encoding='utf-8') as f:
                    json.dump(sqs_db, f, indent=4)
                logger.info(f"Updated SQS for {updates_made} suppliers based on PRTS data.")
            except Exception as e:
                logger.error(f"Failed to save SQS DB: {e}")
                
        # Merge back to active cache
        if hasattr(self, 'databases') and 'supplier_quality' in self.databases:
            self.databases['supplier_quality'] = sqs_db

    def update_demand_intelligence(self):
        """Integrates POS Sales and Transfers to update the Sales Forecasting database."""
        sales = self.scan_cashier_sales()
        transf = self.scan_inventory_transfers()
        if not sales and not transf: return
        f_db = self.databases.get('sales_forecasting', {})
        # BUG 6 FIX: Use actual data period instead of hardcoded 300 days.
        # Each *_cash.xlsx represents ~30 days. Count files to get true span.
        import glob
        num_cash_files = max(1, len(glob.glob(os.path.join(self.data_dir, "*_cash.xlsx"))))
        data_span_days = num_cash_files * 30.0
        for key in (set(sales.keys()) | set(transf.keys())):
            entry = f_db.get(key)
            if entry:
                true_demand = sales.get(key, 0.0) + transf.get(key, {}).get('out', 0.0) - transf.get(key, {}).get('in', 0.0)
                calc_daily = max(0.0, true_demand / data_span_days)
                prev = entry.get('avg_daily_sales', 0.0)
                entry['avg_daily_sales'] = round((prev + calc_daily) / 2, 4) if prev > 0 else round(calc_daily, 4)
        self.databases['sales_forecasting'] = f_db

    def update_supplier_patterns(self):
        """Processes PO files and updates the supplier patterns database."""
        import statistics
        po_hist = self.scan_purchase_orders()
        if not po_hist: return
        p_db = self.databases.get('supplier_patterns', {})
        for supplier, dates in po_hist.items():
            if len(dates) < 2: continue
            gaps = [(dates[i] - dates[i-1]).days for i in range(1, len(dates)) if (dates[i]-dates[i-1]).days > 0]
            if not gaps: continue
            if supplier not in p_db: p_db[supplier] = {'estimated_delivery_days': 7, 'reliability_score': 100}
            p_db[supplier].update({'median_gap_days': int(statistics.median(gaps)), 'average_gap_days': round(statistics.mean(gaps), 1), 'total_orders': len(dates)})
        self.databases['supplier_patterns'] = p_db

    def update_lead_time_intelligence(self):
        """Calculates fulfillment lead times by linking PO dates to GRN receipt dates."""
        import glob
        import statistics
        po_dates = {}
        for fpath in glob.glob(os.path.join(self.data_dir, "po_*.xlsx")):
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                h = {"".join(c for c in str(v).lower() if c.isalnum()): i for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))}
                c_po, c_dt = h.get('pono'), h.get('podate', h.get('date'))
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if c_po is not None and c_dt is not None and row[c_po] and row[c_dt]:
                        po_dates[str(row[c_po]).strip()] = row[c_dt]
                wb.close()
            except: pass
        
        lt_stats = {}
        for fpath in glob.glob(os.path.join(self.data_dir, "grnd*.xlsx")):
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                h = {"".join(c for c in str(v).lower() if c.isalnum()): i for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))}
                c_po, c_dt, c_vn = h.get('pono'), h.get('grndate', h.get('docdate')), h.get('vendorcodename', h.get('vendor'))
                
                if c_po is None or c_dt is None or c_vn is None:
                    logger.debug(f"Skipping GRN {os.path.basename(fpath)}: Missing required columns (PO, Date, Vendor)")
                    continue

                matches = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    po_no = str(row[c_po]).strip()
                    if po_no in po_dates and row[c_dt]:
                        d1_val = row[c_dt]
                        d2_val = po_dates[po_no]
                        
                        def parse_dt(v):
                            from datetime import datetime as dt, date
                            if isinstance(v, (dt, date)): return v
                            if isinstance(v, str):
                                for fmt in ('%d-%b-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                                    try: return dt.strptime(v.strip(), fmt)
                                    except: continue
                            return None

                        d1 = parse_dt(d1_val)
                        d2 = parse_dt(d2_val)
                        
                        from datetime import datetime as dt
                        if d1 and d2:
                            # Standardize to date for subtraction
                            dt1 = d1.date() if isinstance(d1, dt) else d1
                            dt2 = d2.date() if isinstance(d2, dt) else d2
                            
                            gap = (dt1 - dt2).days
                            if gap >= 0:
                                v_raw = str(row[c_vn])
                                supp = v_raw.split(' - ', 1)[1].upper().strip() if ' - ' in v_raw else v_raw.upper().strip()
                                if supp not in lt_stats: lt_stats[supp] = []
                                lt_stats[supp].append(gap)
                                matches += 1
                if matches > 0:
                    logger.info(f"Found {matches} lead-time matches in {os.path.basename(fpath)}")
                wb.close()
            except Exception as e:
                logger.error(f"Error processing GRN {os.path.basename(fpath)}: {e}")
        
        p_db = self.databases.get('supplier_patterns', {})
        for s, gaps in lt_stats.items():
            if s in p_db: p_db[s]['estimated_delivery_days'] = max(1, int(statistics.median(gaps)))
        self.databases['supplier_patterns'] = p_db

        # v1.1 Upgrade: Save raw gaps for LATA precision
        if lt_stats:
            try:
                import json
                gaps_path = os.path.join(self.data_dir, 'supplier_delivery_gaps.json')
                with open(gaps_path, 'w', encoding='utf-8') as f:
                    json.dump(lt_stats, f, indent=4)
                logger.info(f"Saved raw delivery gaps for {len(lt_stats)} suppliers to {os.path.basename(gaps_path)}")
            except Exception as e:
                logger.error(f"Failed to save delivery gaps: {e}")

    def scan_sales_profitability(self):
        """Scans topselqty.xlsx and updates profitability database."""
        fpath = os.path.join(self.data_dir, "topselqty.xlsx")
        if not os.path.exists(fpath): return
        try:
            wb = load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
            h = {str(v).strip().lower(): i for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))}
            c_nm, c_qty, c_rev, c_mpct = h.get('item name'), h.get('qty'), h.get('net amt'), h.get('margin %')
            prof = {}
            rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[c_nm] and isinstance(r[c_qty], (int, float))]
            rows.sort(key=lambda x: x[c_qty], reverse=True)
            for i, r in enumerate(rows, 1):
                name = str(r[c_nm]).strip().upper()
                prof[name] = {"total_qty_sold": r[c_qty], "revenue": float(r[c_rev] or 0), "margin_pct": float(r[c_mpct] or 0), "sales_rank": i}
            self.databases['sales_profitability'] = prof
            wb.close()
        except Exception as e: logger.error(f"Profitability scan error: {e}")
