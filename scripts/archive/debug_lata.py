import os
import glob
import openpyxl
from datetime import datetime

data_dir = r'C:\Users\iLink\.gemini\antigravity\scratch\oasis\data'
po_dates = {}
print(f"Scanning POs in {data_dir}...")
for fpath in glob.glob(os.path.join(data_dir, "po_*.xlsx")):
    print(f"  PO File: {os.path.basename(fpath)}")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    h = {"".join(c for c in str(v).lower() if c.isalnum()): i for i, v in enumerate(header_row) if v}
    print(f"    Headers: {h}")
    c_po, c_dt = h.get('pono'), h.get('podate')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[c_po] and row[c_dt]:
            po_dates[str(row[c_po]).strip()] = row[c_dt]
    wb.close()

print(f"Captured {len(po_dates)} PO dates.")
if len(po_dates) > 0:
    print(f"  Example PO: {list(po_dates.keys())[0]} -> {po_dates[list(po_dates.keys())[0]]}")

lt_stats = {}
print(f"Scanning GRNs in {data_dir}...")
for fpath in glob.glob(os.path.join(data_dir, "grnd*.xlsx")):
    print(f"  GRN File: {os.path.basename(fpath)}")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    h = {"".join(c for c in str(v).lower() if c.isalnum()): i for i, v in enumerate(header_row) if v}
    print(f"    Headers: {h}")
    c_po, c_dt, c_vn = h.get('pono'), h.get('grndate', h.get('docdate')), h.get('vendorcodename', h.get('vendor'))
    
    match_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        po_no = str(row[c_po]).strip()
        if po_no in po_dates:
            match_count += 1
            d1 = row[c_dt]
            d2 = po_dates[po_no]
            if isinstance(d1, datetime) and isinstance(d2, datetime):
                gap = (d1 - d2).days
                if gap >= 0:
                    v_raw = str(row[c_vn])
                    supp = v_raw.split(' - ', 1)[1].upper().strip() if ' - ' in v_raw else v_raw.upper().strip()
                    if supp not in lt_stats: lt_stats[supp] = []
                    lt_stats[supp].append(gap)
    print(f"    Matches found: {match_count}")
    wb.close()

print(f"Final lt_stats: {len(lt_stats)} suppliers.")
