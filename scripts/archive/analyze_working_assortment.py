"""
Rhapta Road Store: Working Assortment Analysis
================================================
1. Loads the dept_*.xlsx stock snapshot files
2. Filters to "working assortment" (stock > 0, excludes consignment fresh & defunct depts)
3. Loads GRN data to get cost prices per barcode
4. Compares working assortment against Allocation Scorecard v7
5. Outputs summary stats and a comparison CSV
"""
import openpyxl
import pandas as pd
import os
import json
from collections import defaultdict

DATA_DIR = os.path.join(os.path.dirname(__file__), "oasis", "data")
BASE_DIR = os.path.dirname(__file__)

# ── Departments to exclude ──────────────────────────────────────────────────
EXCLUDE_DEPTS = {
    # Consignment suppliers (fresh produce counted at POS, not owned stock)
    "FRESH VEG CS", "FRESH FRUITS CS", "FRESH HERBS CS",
    # Defunct departments
    "BAKERY BIGCOLD",
}

# ── 1. Load Department Stock Snapshot ────────────────────────────────────────
print("=" * 80)
print("  STEP 1: Loading Rhapta Store Stock Snapshot")
print("=" * 80)

dept_files = [
    os.path.join(DATA_DIR, "dept_1_50.xlsx"),
    os.path.join(DATA_DIR, "dept_51_100.xlsx"),
    os.path.join(DATA_DIR, "dept_101_150.xlsx"),
    os.path.join(DATA_DIR, "dept_151_200.xlsx"),
    os.path.join(DATA_DIR, "dept_201_250.xlsx"),
    os.path.join(DATA_DIR, "dept_301_350.xlsx"),
]

all_items = []
for f in dept_files:
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        all_items.append(item)
    wb.close()

df_stock = pd.DataFrame(all_items)
df_stock['STOCK'] = pd.to_numeric(df_stock['STOCK'], errors='coerce').fillna(0)
df_stock['SellPrice'] = pd.to_numeric(df_stock['SellPrice'], errors='coerce').fillna(0)
df_stock['BARCODE'] = df_stock['BARCODE'].astype(str).str.strip()
df_stock['ITM_NAME'] = df_stock['ITM_NAME'].astype(str).str.strip()
df_stock['DEPARTMENT'] = df_stock['DEPARTMENT'].astype(str).str.strip()
df_stock['VENDOR_NAME'] = df_stock['VENDOR_NAME'].astype(str).str.strip()
df_stock['StockValue'] = df_stock['STOCK'] * df_stock['SellPrice']

print(f"  Total SKUs loaded:          {len(df_stock):>8,}")
print(f"  Zero-stock items:           {(df_stock['STOCK'] == 0).sum():>8,}")
print(f"  Negative-stock items:       {(df_stock['STOCK'] < 0).sum():>8,}")
print()

# ── 2. Filter to Working Assortment ─────────────────────────────────────────
print("=" * 80)
print("  STEP 2: Filtering to Working Assortment")
print("=" * 80)

# Exclude consignment & defunct
excluded_mask = df_stock['DEPARTMENT'].str.upper().isin({d.upper() for d in EXCLUDE_DEPTS})
excluded_count = excluded_mask.sum()
print(f"  Excluded (consignment/defunct): {excluded_count:>6,} items")
for dept in sorted(EXCLUDE_DEPTS):
    c = (df_stock['DEPARTMENT'].str.upper() == dept.upper()).sum()
    if c > 0:
        print(f"    - {dept}: {c:,}")

# Filter: stock > 0  AND  not in excluded departments
df_working = df_stock[(df_stock['STOCK'] > 0) & (~excluded_mask)].copy()
df_working = df_working.drop_duplicates(subset='BARCODE', keep='first')

print(f"\n  Working assortment (stock > 0, non-excluded):")
print(f"    SKUs:                     {len(df_working):>8,}")
print(f"    Total units:              {df_working['STOCK'].sum():>12,.0f}")
print(f"    Stock value (sell price): KES {df_working['StockValue'].sum():>12,.0f}")
print(f"    Departments:              {df_working['DEPARTMENT'].nunique():>8}")
print(f"    Vendors:                  {df_working['VENDOR_NAME'].nunique():>8}")
print()

# ── 3. Load GRN Data for Cost Prices ────────────────────────────────────────
print("=" * 80)
print("  STEP 3: Loading GRN Data for Cost Prices")
print("=" * 80)

grn_files = [f for f in os.listdir(DATA_DIR) if f.startswith("grnds_") and f.endswith(".xlsx")]
grn_files.append("grnd_1_1.5.xlsx")  # The alternate naming one

cost_lookup = {}  # barcode -> (cost_price, item_name, vendor)
grn_rows = 0
for gf in grn_files:
    path = os.path.join(DATA_DIR, gf)
    if not os.path.exists(path):
        continue
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    bc_idx = headers.index('Bar Code')
    cost_idx = headers.index('Cost Price')
    name_idx = headers.index('Item Name')
    vendor_idx = headers.index('Vendor Code - Name')

    for row in ws.iter_rows(min_row=2, values_only=True):
        barcode = str(row[bc_idx]).strip() if row[bc_idx] else ''
        cost = row[cost_idx] if row[cost_idx] else 0
        name = row[name_idx] if row[name_idx] else ''
        vendor = row[vendor_idx] if row[vendor_idx] else ''
        if barcode and cost > 0:
            # Keep the most recent (last seen) cost price
            cost_lookup[barcode] = (cost, name, vendor)
        grn_rows += 1
    wb.close()

print(f"  GRN files loaded:           {len(grn_files):>8}")
print(f"  GRN line items:             {grn_rows:>8,}")
print(f"  Unique barcodes with cost:  {len(cost_lookup):>8,}")
print()

# ── 4. Enrich Working Assortment with Cost Price ────────────────────────────
print("=" * 80)
print("  STEP 4: Enriching Working Assortment with Cost Prices")
print("=" * 80)

df_working['CostPrice'] = df_working['BARCODE'].map(
    lambda bc: cost_lookup.get(bc, (None,))[0]
)
df_working['CostPrice'] = pd.to_numeric(df_working['CostPrice'], errors='coerce')

has_cost = df_working['CostPrice'].notna()
print(f"  Items with GRN cost price:  {has_cost.sum():>8,} ({has_cost.mean()*100:.1f}%)")
print(f"  Items without cost price:   {(~has_cost).sum():>8,}")

# For items without GRN cost, estimate from sell price using avg margin
avg_margin = 0.25  # Default 25% margin assumption as fallback
df_working['CostPrice_Final'] = df_working['CostPrice'].fillna(
    df_working['SellPrice'] * (1 - avg_margin)
)
df_working['CostValue'] = df_working['STOCK'] * df_working['CostPrice_Final']
df_working['Margin_Pct'] = (
    (df_working['SellPrice'] - df_working['CostPrice_Final']) / df_working['SellPrice'] * 100
).round(1)

print(f"\n  Working assortment value (at cost): KES {df_working['CostValue'].sum():>12,.0f}")
print(f"  Working assortment value (at sell): KES {df_working['StockValue'].sum():>12,.0f}")
print(f"  Implied avg margin:                 {((df_working['StockValue'].sum() - df_working['CostValue'].sum()) / df_working['StockValue'].sum() * 100):.1f}%")
print()

# ── 5. Load Allocation Scorecard ────────────────────────────────────────────
print("=" * 80)
print("  STEP 5: Comparing Against Allocation Scorecard v7")
print("=" * 80)

df_sc = pd.read_csv(os.path.join(BASE_DIR, "Full_Product_Allocation_Scorecard_v7.csv"))
df_sc['Product_upper'] = df_sc['Product'].str.strip().str.upper()
df_sc['Department_upper'] = df_sc['Department'].str.strip().str.upper()

# Join working assortment to scorecard by item name
df_working['ITM_NAME_upper'] = df_working['ITM_NAME'].str.upper()
df_merged = df_working.merge(
    df_sc, left_on='ITM_NAME_upper', right_on='Product_upper',
    how='left', suffixes=('_stock', '_sc')
)

matched = df_merged['Product'].notna()
print(f"  Working items matched to scorecard: {matched.sum():>8,} ({matched.mean()*100:.1f}%)")
print(f"  Working items NOT in scorecard:     {(~matched).sum():>8,}")
print()

# ── 6. Detailed Comparison ──────────────────────────────────────────────────
print("=" * 80)
print("  STEP 6: Working Assortment by Department (Top 25)")
print("=" * 80)

dept_summary = df_working.groupby('DEPARTMENT').agg(
    SKUs=('BARCODE', 'count'),
    TotalUnits=('STOCK', 'sum'),
    SellValue=('StockValue', 'sum'),
    CostValue=('CostValue', 'sum'),
    Vendors=('VENDOR_NAME', 'nunique'),
    AvgMargin=('Margin_Pct', 'mean'),
).sort_values('SellValue', ascending=False)

print(f"{'Department':<25} {'SKUs':>5} {'Units':>8} {'Sell Val':>12} {'Cost Val':>12} {'Margin%':>7} {'Vendors':>7}")
print("-" * 80)
for dept, row in dept_summary.head(25).iterrows():
    print(f"{dept[:24]:<25} {row['SKUs']:>5} {row['TotalUnits']:>8,.0f} {row['SellValue']:>12,.0f} {row['CostValue']:>12,.0f} {row['AvgMargin']:>6.1f}% {row['Vendors']:>7}")

print()

# ── 7. Scorecard vs Reality Gap Analysis ────────────────────────────────────
print("=" * 80)
print("  STEP 7: Scorecard vs Reality - Gap Analysis")
print("=" * 80)

df_match = df_merged[matched].copy()

# Items in scorecard that are ELIGIBLE but have low stock
df_match['SC_Recommended'] = pd.to_numeric(df_match.get('Recommended_Qty', 0), errors='coerce').fillna(0)
df_match['SC_Priority'] = df_match.get('Priority_Label', 'Unknown')
df_match['SC_Velocity'] = df_match.get('Velocity_Tier', 'Unknown')
df_match['SC_ADS'] = pd.to_numeric(df_match.get('Avg_Daily_Sales', 0), errors='coerce').fillna(0)

# Stock coverage in days
df_match['DaysOfCover'] = (df_match['STOCK'] / df_match['SC_ADS'].replace(0, float('nan'))).round(1)

# Gap: Recommended - Actual
df_match['StockGap'] = df_match['SC_Recommended'] - df_match['STOCK']

print("\n  Distribution by Priority Label (matched items):")
priority_dist = df_match.groupby('SC_Priority').agg(
    Count=('BARCODE', 'count'),
    AvgStock=('STOCK', 'mean'),
    AvgRecommended=('SC_Recommended', 'mean'),
    AvgDaysCover=('DaysOfCover', 'mean'),
    TotalSellValue=('StockValue', 'sum'),
).round(1)
print(priority_dist.to_string())

print("\n  Distribution by Velocity Tier (matched items):")
velocity_dist = df_match.groupby('SC_Velocity').agg(
    Count=('BARCODE', 'count'),
    AvgStock=('STOCK', 'mean'),
    AvgRecommended=('SC_Recommended', 'mean'),
    AvgDaysCover=('DaysOfCover', 'mean'),
    TotalCostValue=('CostValue', 'sum'),
).round(1)
print(velocity_dist.to_string())

# ── 8. Items in Scorecard but NOT in working assortment ─────────────────────
print()
print("=" * 80)
print("  STEP 8: Scorecard Items Missing from Working Assortment")
print("=" * 80)

# All eligible scorecard items
sc_eligible = df_sc[df_sc.get('Is_Eligible_Basic', False) == True].copy()
sc_names = set(sc_eligible['Product_upper'])
working_names = set(df_working['ITM_NAME_upper'])

scorecard_only = sc_names - working_names
in_both = sc_names & working_names

print(f"  Scorecard eligible items:       {len(sc_eligible):>8,}")
print(f"  In both (stocked & eligible):   {len(in_both):>8,}")
print(f"  Scorecard-only (not stocked):   {len(scorecard_only):>8,}")
print(f"  Stocked but not in scorecard:   {len(working_names - sc_names):>8,}")

# Top scorecard-only items by revenue
sc_missing = sc_eligible[sc_eligible['Product_upper'].isin(scorecard_only)].copy()
sc_missing_top = sc_missing.nlargest(20, 'Total_Revenue')

print(f"\n  Top 20 Scorecard Items NOT Currently Stocked (by Revenue):")
print(f"  {'Product':<45} {'Dept':<18} {'ADS':>5} {'Revenue':>12} {'Priority':>10}")
print("  " + "-" * 95)
for _, r in sc_missing_top.iterrows():
    print(f"  {str(r['Product'])[:44]:<45} {str(r['Department'])[:17]:<18} {r['Avg_Daily_Sales']:>5.1f} {r['Total_Revenue']:>12,.0f} {str(r.get('Priority_Label','?')):>10}")

# ── 9. Items stocked but NOT in scorecard ───────────────────────────────────
print()
print("=" * 80)
print("  STEP 9: Stocked Items NOT in Scorecard (Top 20 by Stock Value)")
print("=" * 80)

not_in_sc = df_working[~df_working['ITM_NAME_upper'].isin(df_sc['Product_upper'])].copy()
not_in_sc_top = not_in_sc.nlargest(20, 'StockValue')

print(f"  {'Item':<45} {'Dept':<18} {'Stock':>6} {'Sell Val':>10}")
print("  " + "-" * 85)
for _, r in not_in_sc_top.iterrows():
    print(f"  {str(r['ITM_NAME'])[:44]:<45} {str(r['DEPARTMENT'])[:17]:<18} {r['STOCK']:>6,.0f} {r['StockValue']:>10,.0f}")

# ── 10. Save Outputs ────────────────────────────────────────────────────────
print()
print("=" * 80)
print("  STEP 10: Saving Outputs")
print("=" * 80)

# Save working assortment
output_cols = [
    'BARCODE', 'ITM_NAME', 'DEPARTMENT', 'VENDOR_NAME',
    'STOCK', 'SellPrice', 'CostPrice_Final', 'StockValue', 'CostValue', 'Margin_Pct'
]
df_working[output_cols].to_csv(
    os.path.join(BASE_DIR, "rhapta_working_assortment.csv"), index=False
)
print(f"  Saved: rhapta_working_assortment.csv ({len(df_working)} items)")

# Save gap analysis (matched items)
if len(df_match) > 0:
    gap_cols = [
        'BARCODE', 'ITM_NAME', 'DEPARTMENT', 'VENDOR_NAME',
        'STOCK', 'SellPrice', 'CostPrice_Final', 'StockValue', 'CostValue',
        'SC_ADS', 'SC_Recommended', 'StockGap', 'DaysOfCover',
        'SC_Priority', 'SC_Velocity',
    ]
    existing = [c for c in gap_cols if c in df_match.columns]
    df_match[existing].to_csv(
        os.path.join(BASE_DIR, "rhapta_scorecard_comparison.csv"), index=False
    )
    print(f"  Saved: rhapta_scorecard_comparison.csv ({len(df_match)} items)")

print()
print("  DONE.")
