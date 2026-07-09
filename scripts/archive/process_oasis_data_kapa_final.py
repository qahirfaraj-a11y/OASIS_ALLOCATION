import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_portfolio():
    data_dir = r"C:\Users\iLink\.gemini\antigravity\scratch\oasis\data"
    nodes_csv = r"C:\Users\iLink\.gemini\antigravity\scratch\neutral_network_export\nodes.csv"
    edges_csv = r"C:\Users\iLink\.gemini\antigravity\scratch\neutral_network_export\edges.csv"
    kapa_excel_path = r"C:\Users\iLink\Downloads\kapa.xlsx"
    
    grn_files = [
        "grnds_2_2.5.xlsx",
        "grnds_2_3.0.xlsx",
        "grnds_3.5_4.xlsx",
        "grnds_3_3.5.xlsx",
        "grnds_7.5_8.xlsx",
        "grnds_8.5_9.xlsx",
        "grnds_8_8.5.xlsx",
        "grnds_9.5_10.xlsx",
        "grnds_9_9.5.xlsx",
        "grnds_10.5_11.xlsx",
        "grnds_10_10.5.xlsx",
        "grnds_11.5_12.xlsx",
        "grnds_11_11.5.xlsx",
        "grnds_12.xlsx",
        "grnd_1_1.5.xlsx",
        "grnds_1_1.5.xlsx",
        "grnds_1_2.0.xlsx"
    ]
    
    dept_files = [
        "dept_101_150.xlsx",
        "dept_151_200.xlsx",
        "dept_201_250.xlsx",
        "dept_301_350.xlsx",
        "dept_1_50.xlsx",
        "dept_51_100.xlsx"
    ]
    
    # 1. Load Nodes & Edges
    df_nodes = pd.read_csv(nodes_csv)
    df_edges = pd.read_csv(edges_csv)
    df_nodes['id_upper'] = df_nodes['id'].astype(str).str.strip().str.upper()
    
    in_degree = df_edges['target'].value_counts().to_dict()
    out_degree = df_edges['source'].value_counts().to_dict()
    
    # 2. Extract GRN cost data
    grn_rows = []
    print("Loading SKU-level GRN transacted costs...")
    for f in grn_files:
        fp = os.path.join(data_dir, f)
        if os.path.exists(fp):
            try:
                df = pd.read_excel(fp)
                if 'Bar Code' in df.columns and 'Vendor Code - Name' in df.columns:
                    kapa = df[df['Vendor Code - Name'].astype(str).str.contains('KAPA', case=False, na=False)].copy()
                    if not kapa.empty:
                        grn_rows.append(kapa)
            except Exception as e:
                print(f"  Error reading GRN {f}: {e}")
                
    if grn_rows:
        df_grn_kapa = pd.concat(grn_rows, ignore_index=True)
        df_grn_kapa['barcode_clean'] = df_grn_kapa['Bar Code'].astype(str).str.split('.').str[0].str.strip()
        df_grn_kapa['item_upper'] = df_grn_kapa['Item Name'].astype(str).str.strip().str.upper()
        # Calculate CP including VAT (Net Amt / GRN Qty)
        df_grn_kapa['calculated_cp'] = df_grn_kapa['Net Amt'] / df_grn_kapa['GRN Qty']
        df_grn_kapa = df_grn_kapa[df_grn_kapa['calculated_cp'] > 0]
        
        barcode_cp_map = df_grn_kapa.groupby('barcode_clean')['calculated_cp'].mean().to_dict()
        name_cp_map = df_grn_kapa.groupby('item_upper')['calculated_cp'].mean().to_dict()
        print(f"Successfully processed {len(df_grn_kapa)} Kapa GRN rows from new data.")
    else:
        barcode_cp_map = {}
        name_cp_map = {}
        print("No Kapa GRN rows loaded.")
        
    # 3. Extract Department selling price & stock
    dept_rows = []
    print("Loading Department selling prices and stock levels...")
    for f in dept_files:
        fp = os.path.join(data_dir, f)
        if os.path.exists(fp):
            try:
                df = pd.read_excel(fp)
                kapa = df[df['VENDOR_NAME'].astype(str).str.contains('KAPA', case=False, na=False)].copy()
                if not kapa.empty:
                    dept_rows.append(kapa)
            except Exception as e:
                print(f"  Error reading Dept {f}: {e}")
                
    if dept_rows:
        df_dept_kapa = pd.concat(dept_rows, ignore_index=True)
        df_dept_kapa['barcode_clean'] = df_dept_kapa['BARCODE'].astype(str).str.split('.').str[0].str.strip()
        df_dept_kapa['item_upper'] = df_dept_kapa['ITM_NAME'].astype(str).str.strip().str.upper()
        
        barcode_sp_map = df_dept_kapa.set_index('barcode_clean')['SellPrice'].to_dict()
        barcode_stock_map = df_dept_kapa.set_index('barcode_clean')['STOCK'].to_dict()
        name_sp_map = df_dept_kapa.set_index('item_upper')['SellPrice'].to_dict()
        name_stock_map = df_dept_kapa.set_index('item_upper')['STOCK'].to_dict()
        print(f"Successfully processed {len(df_dept_kapa)} Kapa Department rows.")
    else:
        barcode_sp_map = {}
        barcode_stock_map = {}
        name_sp_map = {}
        name_stock_map = {}
        print("No Kapa Department rows loaded.")
        
    # 4. Load Original Catalog to get barcodes and backup prices
    df_kapa_data = pd.read_excel(kapa_excel_path, header=2)
    catalog_barcodes = {}
    catalog_sps = {}
    for idx, row in df_kapa_data.iterrows():
        desc = row['DESCRIPTION']
        if pd.isna(desc):
            continue
        desc_upper = str(desc).strip().upper()
        barcode = str(row['BARCODE']).split('.')[0].strip()
        sp_raw = row['SP']
        if isinstance(sp_raw, str):
            sp = float(sp_raw.replace(',', ''))
        else:
            sp = float(sp_raw)
        catalog_barcodes[desc_upper] = barcode
        catalog_sps[desc_upper] = sp
        
    # 5. Calculate category margins from newly transacted items for imputation fallback
    matched_margins = []
    for desc_upper, barcode in catalog_barcodes.items():
        sp = barcode_sp_map.get(barcode, name_sp_map.get(desc_upper, catalog_sps.get(desc_upper, 0.0)))
        cp = barcode_cp_map.get(barcode, name_cp_map.get(desc_upper, None))
        if cp is not None and cp > 0 and sp > 0:
            margin = (sp - cp) / sp
            # get department from nodes
            match_node = df_nodes[df_nodes['id_upper'] == desc_upper]
            dept = 'GENERAL'
            if not match_node.empty:
                raw_dept = str(match_node.iloc[0]['department'])
                dept = raw_dept.replace('[[','').replace(']]','').strip()
            matched_margins.append({'dept': dept, 'margin': margin})
            
    df_mm = pd.DataFrame(matched_margins)
    if not df_mm.empty:
        dept_avg_margins = df_mm.groupby('dept')['margin'].mean().to_dict()
        overall_avg_margin = df_mm['margin'].mean()
    else:
        dept_avg_margins = {}
        overall_avg_margin = 0.2840
    print(f"Overall transacted average margin: {overall_avg_margin * 100:.2f}%")
    
    # 6. Build Sheet 2: Kapa Network Nodes
    kapa_nodes = df_nodes[df_nodes['supplier'].astype(str).str.contains('KAPA', case=False, na=False)].copy()
    kapa_nodes['department_clean'] = kapa_nodes['department'].astype(str).str.replace('[[','', regex=False).str.replace(']]','', regex=False).str.strip()
    
    node_rows = []
    for _, row in kapa_nodes.iterrows():
        sku_id = row['id_upper']
        # Try to find barcode in catalog matching
        barcode = catalog_barcodes.get(sku_id, '')
        
        # Get Sell Price SP
        sp = barcode_sp_map.get(barcode, name_sp_map.get(sku_id, None))
        if sp is None or sp <= 0:
            # PRIORITIZE catalog price to eliminate GAT derived pricing anomalies
            sp = catalog_sps.get(sku_id, None)
            if sp is None or sp <= 0:
                sp = float(row['price'])
            
        # Get Cost Price CP
        cp = barcode_cp_map.get(barcode, name_cp_map.get(sku_id, None))
        is_imputed = False
        if cp is None or cp <= 0:
            dept = row['department_clean']
            margin = dept_avg_margins.get(dept, overall_avg_margin)
            cp = sp * (1 - margin)
            is_imputed = True
            
        # Get stock level
        stock = barcode_stock_map.get(barcode, name_stock_map.get(sku_id, 0.0))
        if pd.isna(stock) or stock < 0:
            stock = 0.0
            
        # Calculations
        ads = float(row['velocity_ads'])
        in_d = in_degree.get(row['id'], 0)
        out_d = out_degree.get(row['id'], 0)
        
        # Periodic Review inventory formulas (T = 13d, L = 5d -> Holding Days = 11.5)
        if ads > 0 and cp > 0:
            roi = (30 * (sp - cp)) / (11.5 * cp) * 100
        else:
            roi = 0.0
            
        safety_stock = 5.0 * ads
        rop = 5.0 * ads
        tsl = 18.0 * ads
        avg_inv_cost = 11.5 * ads * cp
        gp_30d = 30 * ads * (sp - cp)
        stock_val = stock * cp
        
        node_rows.append({
            'SKU Name': row['id'],
            'Department': row['department_clean'],
            'Selling Price (SP)': sp,
            'Cost Price (CP)': cp,
            'Is Cost Imputed': 'Yes' if is_imputed else 'No',
            'Gross Margin (%)': ((sp - cp) / sp * 100) if sp > 0 else 0.0,
            'Daily Sales Velocity (ADS)': ads,
            'Current Stock (Units)': stock,
            'Stock Valuation (KES)': stock_val,
            'Avg Inventory (Units)': 11.5 * ads,
            'Avg Inventory Cost (KES)': avg_inv_cost,
            '30D Gross Profit (KES)': gp_30d,
            '30D ROI (%)': roi,
            'Safety Stock (SS)': safety_stock,
            'Target Stock Level (TSL)': tsl,
            'Attractor Score': in_d,
            'Connector Score': out_d,
            'Sales Rank': row['sales_rank'],
            'Fill Rate (%)': row['rhapta_fill_rate'] * 100
        })
    df_nodes_out = pd.DataFrame(node_rows)
    df_nodes_out.sort_values(by='30D ROI (%)', ascending=False, inplace=True)
    
    # 7. Build Sheet 3: Catalog Audit
    catalog_rows = []
    for desc_upper, barcode in catalog_barcodes.items():
        # Find match in nodes
        match = df_nodes[df_nodes['id_upper'] == desc_upper]
        if match.empty:
            match = df_nodes[df_nodes['id_upper'].str.contains(desc_upper, regex=False)]
            
        # Get SP, CP, Stock
        sp = barcode_sp_map.get(barcode, name_sp_map.get(desc_upper, catalog_sps.get(desc_upper, 0.0)))
        if sp is None or sp <= 0:
            sp = catalog_sps.get(desc_upper, 0.0)
            
        cp = barcode_cp_map.get(barcode, name_cp_map.get(desc_upper, None))
        is_imputed = False
        if cp is None or cp <= 0:
            if not match.empty:
                raw_dept = str(match.iloc[0]['department'])
                dept = raw_dept.replace('[[','').replace(']]','').strip()
            else:
                dept = 'GENERAL'
            margin = dept_avg_margins.get(dept, overall_avg_margin)
            cp = sp * (1 - margin)
            is_imputed = True
            
        stock = barcode_stock_map.get(barcode, name_stock_map.get(desc_upper, 0.0))
        if pd.isna(stock) or stock < 0:
            stock = 0.0
            
        if not match.empty:
            row = match.iloc[0]
            ads = float(row['velocity_ads'])
            in_d = in_degree.get(row['id'], 0)
            out_d = out_degree.get(row['id'], 0)
            status = 'Mapped'
        else:
            ads = 0.0
            in_d = 0
            out_d = 0
            status = 'Not in Network'
            
        roi = (30 * (sp - cp)) / (11.5 * cp) * 100 if (ads > 0 and cp > 0) else 0.0
        stock_val = stock * cp
        
        catalog_rows.append({
            'Catalog Item Name': desc_upper,
            'Network Mapping Status': status,
            'Selling Price (SP)': sp,
            'Cost Price (CP)': cp,
            'Current Stock (Units)': stock,
            'Stock Valuation (KES)': stock_val,
            'Derived Margin (%)': ((sp - cp) / sp * 100) if sp > 0 else 0.0,
            'Attractor Score': in_d,
            'Connector Score': out_d,
            'Derived ROI (%)': roi
        })
    df_catalog_out = pd.DataFrame(catalog_rows)
    df_catalog_out.sort_values(by='Derived ROI (%)', ascending=False, inplace=True)
    
    # 8. Save Workbook and Format using Openpyxl
    out_paths = [
        r"C:\Users\iLink\Downloads\Kapa_Portfolio_Node_Intelligence.xlsx",
        r"C:\Users\iLink\.gemini\antigravity\scratch\Kapa_Portfolio_Node_Intelligence.xlsx"
    ]
    
    for out_path in out_paths:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary Sheet
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Colors
        navy_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
        font_section = Font(name="Calibri", size=13, bold=True, color="1F497D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_normal = Font(name="Calibri", size=11)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Title
        ws_sum.cell(row=2, column=2, value="Kapa Oil Refineries - Node & Portfolio Intelligence (OASIS Live Update)").font = font_title
        
        # Summary Statistics
        ws_sum.cell(row=4, column=2, value="KPI Metrics").font = font_section
        ws_sum.cell(row=5, column=2, value="Metric").font = font_header
        ws_sum.cell(row=5, column=2).fill = navy_header_fill
        ws_sum.cell(row=5, column=3, value="Value").font = font_header
        ws_sum.cell(row=5, column=3).fill = navy_header_fill
        ws_sum.cell(row=5, column=3).alignment = align_center
        
        kpis = [
            ("Total Catalog SKUs", len(df_catalog_out)),
            ("Mapped Network SKUs", len(df_catalog_out[df_catalog_out['Network Mapping Status'] == 'Mapped'])),
            ("Total Portfolio Nodes in ST-GAT", len(df_nodes_out)),
            ("Average Live transacted margin %", f"{df_nodes_out['Gross Margin (%)'].mean():.2f}%"),
            ("Maximum SKU ROI (30D)", f"{df_nodes_out['30D ROI (%)'].max():.2f}%"),
            ("Average SKU ROI (30D)", f"{df_nodes_out['30D ROI (%)'].mean():.2f}%"),
            ("Total Daily Sales Velocity (ADS)", f"{df_nodes_out['Daily Sales Velocity (ADS)'].sum():.2f} units/day"),
            ("Total Current Stock on Hand (Units)", f"{df_nodes_out['Current Stock (Units)'].sum():,.0f} units"),
            ("Total Current Stock Valuation", f"KES {df_nodes_out['Stock Valuation (KES)'].sum():,.2f}"),
            ("Total Active Inventory Investment Target", f"KES {df_nodes_out['Avg Inventory Cost (KES)'].sum():,.2f}"),
            ("Estimated Monthly Gross Profit Contribution", f"KES {df_nodes_out['30D Gross Profit (KES)'].sum():,.2f}")
        ]
        
        for idx, (m, v) in enumerate(kpis):
            r = 6 + idx
            ws_sum.cell(row=r, column=2, value=m).font = font_normal
            ws_sum.cell(row=r, column=2).border = border_thin
            ws_sum.cell(row=r, column=3, value=v).font = font_bold
            ws_sum.cell(row=r, column=3).border = border_thin
            ws_sum.cell(row=r, column=3).alignment = align_right
            if idx % 2 == 1:
                ws_sum.cell(row=r, column=2).fill = light_gray_fill
                ws_sum.cell(row=r, column=3).fill = light_gray_fill
                
        # Sheet 2: Kapa Network Nodes
        ws_nodes = wb.create_sheet(title="Kapa Network Nodes")
        ws_nodes.views.sheetView[0].showGridLines = True
        
        headers_nodes = list(df_nodes_out.columns)
        for col_idx, h in enumerate(headers_nodes):
            cell = ws_nodes.cell(row=1, column=col_idx+1, value=h)
            cell.font = font_header
            cell.fill = navy_header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        for r_idx, row in enumerate(df_nodes_out.values):
            r = r_idx + 2
            for col_idx, val in enumerate(row):
                cell = ws_nodes.cell(row=r, column=col_idx+1, value=val)
                cell.font = font_normal
                cell.border = border_thin
                
                h = headers_nodes[col_idx]
                if 'Price' in h or 'Cost' in h or 'Profit' in h or 'Valuation' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif 'Margin' in h or 'Rate' in h or 'ROI' in h:
                    cell.number_format = '0.00"%"'
                    cell.alignment = align_right
                elif 'Velocity' in h or 'Units' in h or 'Stock' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                    if 'Stock (Units)' in h and val > 0:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                elif 'Score' in h or 'Rank' in h:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
                    
        # Sheet 3: Catalog Audit
        ws_cat = wb.create_sheet(title="Catalog Audit")
        ws_cat.views.sheetView[0].showGridLines = True
        
        headers_cat = list(df_catalog_out.columns)
        for col_idx, h in enumerate(headers_cat):
            cell = ws_cat.cell(row=1, column=col_idx+1, value=h)
            cell.font = font_header
            cell.fill = navy_header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        for r_idx, row in enumerate(df_catalog_out.values):
            r = r_idx + 2
            for col_idx, val in enumerate(row):
                cell = ws_cat.cell(row=r, column=col_idx+1, value=val)
                cell.font = font_normal
                cell.border = border_thin
                
                h = headers_cat[col_idx]
                if 'Price' in h or 'Cost' in h or 'Valuation' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif 'Margin' in h or 'ROI' in h:
                    cell.number_format = '0.00"%"'
                    cell.alignment = align_right
                elif 'Score' in h or 'Stock' in h:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
                    
                if val == 'Not in Network':
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
                    cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
                    
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(out_path)
        print(f"Workbook successfully updated and saved to: {out_path}")

if __name__ == "__main__":
    build_portfolio()
