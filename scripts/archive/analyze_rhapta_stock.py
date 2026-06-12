"""Analyze Rhapta store stock snapshot from department Excel files."""
import openpyxl
from collections import defaultdict
import os

DATA_DIR = os.path.join(os.path.dirname(__file__), "oasis", "data")

files = [
    os.path.join(DATA_DIR, "dept_1_50.xlsx"),
    os.path.join(DATA_DIR, "dept_51_100.xlsx"),
    os.path.join(DATA_DIR, "dept_101_150.xlsx"),
    os.path.join(DATA_DIR, "dept_151_200.xlsx"),
    os.path.join(DATA_DIR, "dept_201_250.xlsx"),
    os.path.join(DATA_DIR, "dept_301_350.xlsx"),
]

dept_stats = defaultdict(lambda: {
    'items': 0, 'stock_units': 0, 'stock_value': 0,
    'zero_stock': 0, 'negative_stock': 0, 'vendors': set()
})
grand_items = 0
grand_value = 0.0
grand_zero = 0
grand_neg = 0
grand_units = 0.0

for f in files:
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    dept_idx = headers.index('DEPARTMENT')
    stock_idx = headers.index('STOCK')
    price_idx = headers.index('SellPrice')
    vendor_idx = headers.index('VENDOR_NAME')

    for row in ws.iter_rows(min_row=2, values_only=True):
        dept = row[dept_idx]
        stock = row[stock_idx] if row[stock_idx] else 0
        price = row[price_idx] if row[price_idx] else 0
        vendor = row[vendor_idx] or 'Unknown'

        d = dept_stats[dept]
        d['items'] += 1
        d['stock_units'] += stock
        d['stock_value'] += stock * price
        d['vendors'].add(vendor)
        if stock == 0:
            d['zero_stock'] += 1
            grand_zero += 1
        if stock < 0:
            d['negative_stock'] += 1
            grand_neg += 1
        grand_items += 1
        grand_value += stock * price
        grand_units += stock
    wb.close()

print("=" * 60)
print("  RHAPTA ROAD STORE - STOCK SNAPSHOT ANALYSIS")
print("=" * 60)
print(f"  Total SKUs:              {grand_items:>10,}")
print(f"  Total Stock (units):     {grand_units:>10,.0f}")
print(f"  Total Stock Value:       KES {grand_value:>12,.0f}")
print(f"  Zero-stock items:        {grand_zero:>10,} ({grand_zero/grand_items*100:.1f}%)")
print(f"  Negative-stock items:    {grand_neg:>10,}")
print(f"  Departments:             {len(dept_stats):>10}")
print(f"  Unique Vendors:          {len(set().union(*(s['vendors'] for s in dept_stats.values()))):>10}")
print()

# Top 20 departments by stock value
print("=" * 90)
print("  TOP 20 DEPARTMENTS BY STOCK VALUE")
print("=" * 90)
header = f"{'Dept':<6} {'Items':>6} {'Units':>10} {'Value (KES)':>15} {'Zero%':>7} {'Vendors':>7}"
print(header)
print("-" * 90)
sorted_depts = sorted(dept_stats.items(), key=lambda x: x[1]['stock_value'], reverse=True)
for dept, s in sorted_depts[:20]:
    zero_pct = s['zero_stock'] / s['items'] * 100 if s['items'] > 0 else 0
    print(f"{str(dept):<6} {s['items']:>6,} {s['stock_units']:>10,.0f} {s['stock_value']:>15,.0f} {zero_pct:>6.1f}% {len(s['vendors']):>7}")

print()

# Departments with highest zero-stock percentage
print("=" * 70)
print("  TOP 10 DEPARTMENTS BY ZERO-STOCK % (min 20 items)")
print("=" * 70)
zero_depts = [(d, s) for d, s in dept_stats.items() if s['items'] >= 20]
zero_depts.sort(key=lambda x: x[1]['zero_stock'] / x[1]['items'], reverse=True)
for dept, s in zero_depts[:10]:
    zero_pct = s['zero_stock'] / s['items'] * 100
    print(f"  Dept {dept}: {s['zero_stock']}/{s['items']} ({zero_pct:.0f}%) zero stock, Value: KES {s['stock_value']:,.0f}")

print()

# Sample items from top department
print("=" * 90)
print("  SAMPLE: TOP 5 ITEMS BY STOCK VALUE (Top Department)")
print("=" * 90)
top_dept = sorted_depts[0][0]
print(f"  Department: {top_dept}")
# Re-read to get item detail
items = []
for f in files:
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    dept_idx = headers.index('DEPARTMENT')
    stock_idx = headers.index('STOCK')
    price_idx = headers.index('SellPrice')
    name_idx = headers.index('ITM_NAME')
    vendor_idx = headers.index('VENDOR_NAME')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[dept_idx] == top_dept:
            stock = row[stock_idx] if row[stock_idx] else 0
            price = row[price_idx] if row[price_idx] else 0
            items.append((row[name_idx], row[vendor_idx], stock, price, stock * price))
    wb.close()

items.sort(key=lambda x: x[4], reverse=True)
for name, vendor, stock, price, val in items[:5]:
    print(f"  {name[:45]:<45} | {vendor[:25]:<25} | Stock: {stock:>6,.0f} | KES {val:>10,.0f}")

# Also show negative stock items
print()
print("=" * 70)
print("  NEGATIVE STOCK ITEMS (sample)")
print("=" * 70)
neg_items = []
for f in files:
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    dept_idx = headers.index('DEPARTMENT')
    stock_idx = headers.index('STOCK')
    price_idx = headers.index('SellPrice')
    name_idx = headers.index('ITM_NAME')
    for row in ws.iter_rows(min_row=2, values_only=True):
        stock = row[stock_idx] if row[stock_idx] else 0
        if stock < 0:
            price = row[price_idx] if row[price_idx] else 0
            neg_items.append((row[name_idx], row[dept_idx], stock, price))
    wb.close()

neg_items.sort(key=lambda x: x[2])
for name, dept, stock, price in neg_items[:10]:
    print(f"  {name[:45]:<45} | Dept {dept} | Stock: {stock:>6,.0f} | Price: {price:>8,.0f}")
