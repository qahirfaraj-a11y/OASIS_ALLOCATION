import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_excel():
    # Paths
    kapa_excel_path = r"C:\Users\iLink\Downloads\kapa.xlsx"
    detail_path = r"C:\Users\iLink\.gemini\antigravity\scratch\All_Suppliers_Fulfillment_Detail.xlsx"
    nodes_csv = r"C:\Users\iLink\.gemini\antigravity\scratch\neutral_network_export\nodes.csv"
    edges_csv = r"C:\Users\iLink\.gemini\antigravity\scratch\neutral_network_export\edges.csv"
    
    # 1. Load Data
    df_kapa_data = pd.read_excel(kapa_excel_path, header=2)
    master_skus = df_kapa_data['DESCRIPTION'].dropna().astype(str).str.strip().str.upper().tolist()
    
    df_detail = pd.read_excel(detail_path)
    df_nodes = pd.read_csv(nodes_csv)
    df_edges = pd.read_csv(edges_csv)
    
    # 2. Prepare mappings
    kapa_detail = df_detail[df_detail['Vendor Name'].astype(str).str.contains('KAPA', case=False, na=False)].copy()
    kapa_detail['Item_Name_upper'] = kapa_detail['Item Name'].astype(str).str.strip().str.upper()
    kapa_detail['calculated_cp'] = kapa_detail['Net Amt'] / kapa_detail['GRN Qty']
    cp_map = kapa_detail.groupby('Item_Name_upper')['calculated_cp'].mean().to_dict()
    
    in_degree = df_edges['target'].value_counts().to_dict()
    out_degree = df_edges['source'].value_counts().to_dict()
    
    df_nodes['id_upper'] = df_nodes['id'].astype(str).str.strip().str.upper()
    
    # 3. Identify all 109 KAPA Nodes
    kapa_nodes = df_nodes[df_nodes['supplier'].astype(str).str.contains('KAPA', case=False, na=False)].copy()
    
    # Clean supplier and department strings
    kapa_nodes['department_clean'] = kapa_nodes['department'].astype(str).str.replace('[[','', regex=False).str.replace(']]','', regex=False).str.strip()
    kapa_nodes['supplier_clean'] = kapa_nodes['supplier'].astype(str).str.replace('[[','', regex=False).str.replace(']]','', regex=False).str.strip()
    
    # Calculate category margins from GRN transacted items
    matched_margins = []
    for idx, row in df_kapa_data.iterrows():
        desc = row['DESCRIPTION']
        if pd.isna(desc):
            continue
        desc_upper = str(desc).strip().upper()
        sp = row['SP']
        if isinstance(sp, str):
            sp = float(sp.replace(',', ''))
        else:
            sp = float(sp)
            
        cp = cp_map.get(desc_upper, None)
        if cp is not None and cp > 0 and sp > 0:
            margin = (sp - cp) / sp
            # find department of this item if exists in nodes
            match_node = df_nodes[df_nodes['id_upper'] == desc_upper]
            dept = 'GENERAL'
            if not match_node.empty:
                raw_dept = str(match_node.iloc[0]['department'])
                dept = raw_dept.replace('[[','').replace(']]','').strip()
            matched_margins.append({'dept': dept, 'margin': margin})
            
    df_mm = pd.DataFrame(matched_margins)
    if not df_mm.empty:
        dept_avg_margins = df_mm.groupby('dept')['margin'].mean().to_dict()
        overall_avg_margin = df_mm['margin'].mean()
    else:
        dept_avg_margins = {}
        overall_avg_margin = 0.2856 # derived earlier
        
    # 4. Populate Sheet 2: Kapa Network Nodes data
    node_rows = []
    for _, row in kapa_nodes.iterrows():
        sku_id = row['id_upper']
        sp = float(row['price'])
        
        # Get Cost Price (CP)
        cp = cp_map.get(sku_id, None)
        is_imputed = False
        if cp is None or cp <= 0:
            dept = row['department_clean']
            margin = dept_avg_margins.get(dept, overall_avg_margin)
            cp = sp * (1 - margin)
            is_imputed = True
            
        # Get network scores
        in_d = in_degree.get(row['id'], 0)
        out_d = out_degree.get(row['id'], 0)
        
        # Calculations
        ads = float(row['velocity_ads'])
        
        # ROI Formula: T = 13.0 days, L = 5.0 days -> Holding Days = 11.5
        if ads > 0 and cp > 0:
            roi = (30 * (sp - cp)) / (11.5 * cp) * 100
        else:
            roi = 0.0
            
        # Inventory metrics
        safety_stock = 5.0 * ads
        rop = 5.0 * ads # assuming lead time replenishment level
        tsl = 18.0 * ads
        avg_inv_cost = 11.5 * ads * cp
        gp_30d = 30 * ads * (sp - cp)
        
        node_rows.append({
            'SKU Name': row['id'],
            'Department': row['department_clean'],
            'Selling Price (SP)': sp,
            'Cost Price (CP)': cp,
            'Is Cost Imputed': 'Yes' if is_imputed else 'No',
            'Gross Margin (%)': ((sp - cp) / sp * 100) if sp > 0 else 0.0,
            'Daily Sales Velocity (ADS)': ads,
            'Avg Inventory (Units)': 11.5 * ads,
            'Avg Inventory Cost (KES)': avg_inv_cost,
            '30D Gross Profit (KES)': gp_30d,
            '30D ROI (%)': roi,
            'Safety Stock (SS)': safety_stock,
            'Target Stock Level (TSL)': tsl,
            'Attractor Score': in_d,
            'Connector Score': out_d,
            'Sales Rank': row['sales_rank'],
            'Fill Rate (%)': row['rhapta_fill_rate'] * 100
        })
        
    df_nodes_out = pd.DataFrame(node_rows)
    df_nodes_out.sort_values(by='30D ROI (%)', ascending=False, inplace=True)
    
    # 5. Populate Sheet 3: Catalog Matching & Audit
    catalog_rows = []
    for sku_name in master_skus:
        # Find exact or substring match in nodes
        match = df_nodes[df_nodes['id_upper'] == sku_name]
        if match.empty:
            match = df_nodes[df_nodes['id_upper'].str.contains(sku_name, regex=False)]
            
        if not match.empty:
            row = match.iloc[0]
            sku_id = row['id_upper']
            sp = float(row['price'])
            cp = cp_map.get(sku_id, None)
            is_imputed = False
            if cp is None or cp <= 0:
                raw_dept = str(row['department'])
                dept = raw_dept.replace('[[','').replace(']]','').strip()
                margin = dept_avg_margins.get(dept, overall_avg_margin)
                cp = sp * (1 - margin)
                is_imputed = True
                
            ads = float(row['velocity_ads'])
            roi = (30 * (sp - cp)) / (11.5 * cp) * 100 if (ads > 0 and cp > 0) else 0.0
            in_d = in_degree.get(row['id'], 0)
            out_d = out_degree.get(row['id'], 0)
            status = 'Mapped'
        else:
            # Not found in network
            # Find in raw kapa if possible to get SP
            row_kapa = df_kapa_data[df_kapa_data['DESCRIPTION'].astype(str).str.strip().str.upper() == sku_name]
            sp = 0.0
            if not row_kapa.empty:
                sp_raw = row_kapa.iloc[0]['SP']
                sp = float(str(sp_raw).replace(',', '')) if isinstance(sp_raw, str) else float(sp_raw)
            margin = overall_avg_margin
            cp = sp * (1 - margin)
            roi = 0.0
            in_d = 0
            out_d = 0
            status = 'Not in Network'
            
        catalog_rows.append({
            'Catalog Item Name': sku_name,
            'Network Mapping Status': status,
            'Selling Price (SP)': sp,
            'Cost Price (CP)': cp,
            'Derived Margin (%)': ((sp - cp) / sp * 100) if sp > 0 else 0.0,
            'Attractor Score': in_d,
            'Connector Score': out_d,
            'Derived ROI (%)': roi
        })
    df_catalog_out = pd.DataFrame(catalog_rows)
    df_catalog_out.sort_values(by='Derived ROI (%)', ascending=False, inplace=True)
    
    # 6. Save Workbook and Format using Openpyxl
    out_paths = [
        r"C:\Users\iLink\.gemini\antigravity\scratch\Kapa_Portfolio_Node_Intelligence.xlsx",
        r"C:\Users\iLink\Downloads\Kapa_Portfolio_Node_Intelligence.xlsx"
    ]
    
    for out_path in out_paths:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary Sheet
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Colors
        navy_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
        font_section = Font(name="Calibri", size=13, bold=True, color="1F497D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_normal = Font(name="Calibri", size=11)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        border_thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Add Title
        ws_sum.cell(row=2, column=2, value="Kapa Oil Refineries - Node & Portfolio Intelligence").font = font_title
        
        # Summary Statistics
        ws_sum.cell(row=4, column=2, value="KPI Metrics").font = font_section
        ws_sum.cell(row=5, column=2, value="Metric").font = font_header
        ws_sum.cell(row=5, column=2).fill = navy_header_fill
        ws_sum.cell(row=5, column=3, value="Value").font = font_header
        ws_sum.cell(row=5, column=3).fill = navy_header_fill
        ws_sum.cell(row=5, column=3).alignment = align_center
        
        kpis = [
            ("Total Catalog SKUs", len(df_catalog_out)),
            ("Mapped Network SKUs", len(df_catalog_out[df_catalog_out['Network Mapping Status'] == 'Mapped'])),
            ("Total Portfolio Nodes in ST-GAT", len(df_nodes_out)),
            ("Average transacted margin %", f"{df_nodes_out['Gross Margin (%)'].mean():.2f}%"),
            ("Maximum SKU ROI (30D)", f"{df_nodes_out['30D ROI (%)'].max():.2f}%"),
            ("Average SKU ROI (30D)", f"{df_nodes_out['30D ROI (%)'].mean():.2f}%"),
            ("Total Daily Sales Velocity (ADS)", f"{df_nodes_out['Daily Sales Velocity (ADS)'].sum():.2f} units/day"),
            ("Total Active Inventory Investment", f"KES {df_nodes_out['Avg Inventory Cost (KES)'].sum():,.2f}"),
            ("Estimated Monthly Gross Profit Contribution", f"KES {df_nodes_out['30D Gross Profit (KES)'].sum():,.2f}")
        ]
        
        for idx, (m, v) in enumerate(kpis):
            r = 6 + idx
            ws_sum.cell(row=r, column=2, value=m).font = font_normal
            ws_sum.cell(row=r, column=2).border = border_thin
            ws_sum.cell(row=r, column=3, value=v).font = font_bold
            ws_sum.cell(row=r, column=3).border = border_thin
            ws_sum.cell(row=r, column=3).alignment = align_right
            if idx % 2 == 1:
                ws_sum.cell(row=r, column=2).fill = light_gray_fill
                ws_sum.cell(row=r, column=3).fill = light_gray_fill
                
        # Sheet 2: Kapa Network Nodes
        ws_nodes = wb.create_sheet(title="Kapa Network Nodes")
        ws_nodes.views.sheetView[0].showGridLines = True
        
        # Write headers
        headers_nodes = list(df_nodes_out.columns)
        for col_idx, h in enumerate(headers_nodes):
            cell = ws_nodes.cell(row=1, column=col_idx+1, value=h)
            cell.font = font_header
            cell.fill = navy_header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        # Write data
        for r_idx, row in enumerate(df_nodes_out.values):
            r = r_idx + 2
            for col_idx, val in enumerate(row):
                cell = ws_nodes.cell(row=r, column=col_idx+1, value=val)
                cell.font = font_normal
                cell.border = border_thin
                
                # Format specific columns
                h = headers_nodes[col_idx]
                if 'Price' in h or 'Cost' in h or 'Profit' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif 'Margin' in h or 'Rate' in h or 'ROI' in h:
                    cell.number_format = '0.00"%"'
                    cell.alignment = align_right
                elif 'Velocity' in h or 'Units' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif 'Score' in h or 'Rank' in h:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
                    
        # Sheet 3: Catalog Matching & Audit
        ws_cat = wb.create_sheet(title="Catalog Audit")
        ws_cat.views.sheetView[0].showGridLines = True
        
        # Write headers
        headers_cat = list(df_catalog_out.columns)
        for col_idx, h in enumerate(headers_cat):
            cell = ws_cat.cell(row=1, column=col_idx+1, value=h)
            cell.font = font_header
            cell.fill = navy_header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        # Write data
        for r_idx, row in enumerate(df_catalog_out.values):
            r = r_idx + 2
            for col_idx, val in enumerate(row):
                cell = ws_cat.cell(row=r, column=col_idx+1, value=val)
                cell.font = font_normal
                cell.border = border_thin
                
                # Format specific columns
                h = headers_cat[col_idx]
                if 'Price' in h or 'Cost' in h:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif 'Margin' in h or 'ROI' in h:
                    cell.number_format = '0.00"%"'
                    cell.alignment = align_right
                elif 'Score' in h:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
                    
                # Highlight "Not in Network" in light red
                if val == 'Not in Network':
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
                    cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
                    
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        # handle formatting in length estimation
                        val_str = str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(out_path)
        print(f"Workbook successfully saved to: {out_path}")

if __name__ == "__main__":
    build_excel()
