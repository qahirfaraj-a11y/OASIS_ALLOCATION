from openpyxl import load_workbook
import os

files = ['c:/Users/iLink/.gemini/antigravity/scratch/app/data/trn_1_12.xlsx', 
         'c:/Users/iLink/.gemini/antigravity/scratch/app/data/trout_1_12.xlsx']

for fpath in files:
    if os.path.exists(fpath):
        print(f"\n--- Checking {os.path.basename(fpath)} ---")
        wb = load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(f"Row {i+1}: {row}")
    else:
        print(f"File not found: {fpath}")
