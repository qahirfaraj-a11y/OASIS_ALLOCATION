from openpyxl import load_workbook
import os

fpath = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data/prts_2.xlsx'
if os.path.exists(fpath):
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print(f"Headers: {header_row}")
else:
    print(f"File not found: {fpath}")
