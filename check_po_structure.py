from openpyxl import load_workbook
import os

fpath = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data/po_1-2.xlsx'
if os.path.exists(fpath):
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    print(f"--- Checking {os.path.basename(fpath)} ---")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(f"Row {i+1}: {row}")
else:
    print(f"File not found: {fpath}")
