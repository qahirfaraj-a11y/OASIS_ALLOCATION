import openpyxl
from openpyxl.styles import Font

def create_sample_picking_list(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Picking List"
    
    # Row 1: Supplier (Col 7)
    ws.cell(row=1, column=7).value = "BROOKSIDE DAIRY LIMITED"
    
    # Row 3: Headers
    headers = ["DESCRIPTION", "ITEM CODE", "BARCODE", "RHAPTA", "RR PREV", "RR GRN", "RR PB", "PACK", "SP"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = h
        cell.font = Font(bold=True)
        
    # Row 4+: Data
    data = [
        ["FRESH MILK 500ML", "1001", "6161100001", 50, 1500, 2, 0, 12, 60.0], # Fast mover
        ["YOGURT STRAWBERRY 250ML", "1002", "6161100002", 10, 300, 250, 0, 24, 45.0], # Slow mover (>200 days)
        ["BUTTER 250G", "1003", "6161100003", 5, 100, 10, 0, 20, 250.0], # Regular mover
        ["WHIPPING CREAM 1L", "1004", "6161100004", 0, 50, 5, 1, 10, 800.0] # Blocked
    ]
    
    for r_idx, row in enumerate(data, 4):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx).value = val
            
    wb.save(filename)
    print(f"Sample picking list created: {filename}")

if __name__ == "__main__":
    create_sample_picking_list("sample_picking_list.xlsx")
