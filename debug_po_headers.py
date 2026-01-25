from openpyxl import load_workbook
import os

fpath = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data/po_1-2.xlsx'
wb = load_workbook(fpath, read_only=True, data_only=True)
ws = wb.active
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
headers = {str(val).strip().lower().replace(' ', ''): idx for idx, val in enumerate(header_row) if val}
print(f"Normalized Headers: {headers}")
