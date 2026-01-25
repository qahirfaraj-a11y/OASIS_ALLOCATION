
import os
import calendar
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from app.logic.calendar_analyzer import CalendarAnalyzer

def generate_excel_calendar(output_path: str, year: int = 2026):
    """Generate the Supplier Order Calendar Excel file."""
    
    # 1. Get Data
    analyzer = CalendarAnalyzer()
    analyzer.load_data()
    supplier_data = analyzer.analyze()
    
    # Filter active suppliers
    active_suppliers = {k: v for k, v in supplier_data.items() if v['order_count'] >= 3}
    
    # Categorize
    daily_suppliers = []
    weekly_suppliers = {i: [] for i in range(7)} # 0=Mon, 6=Sun
    monthly_suppliers = {i: [] for i in range(1, 32)} # 1-31
    irregular_suppliers = []
    
    for name, data in active_suppliers.items():
        cat = data['category']
        if cat == 'Daily':
            daily_suppliers.append((name, data['lead_time_days']))
        elif cat == 'Weekly':
            day = data['preferred_day']
            weekly_suppliers[day].append(name)
        elif cat == 'Monthly':
            day = data['preferred_day']
            monthly_suppliers[day].append(name)
        else:
            irregular_suppliers.append((name, data['category']))
            
    # Sort lists
    daily_suppliers.sort(key=lambda x: x[0])
    for d in weekly_suppliers: weekly_suppliers[d].sort()
    for d in monthly_suppliers: monthly_suppliers[d].sort()
    
    # 2. Create Workbook
    wb = Workbook()
    
    # --- SHEET 1: 2026 Order Schedule ---
    ws_schedule = wb.active
    ws_schedule.title = f"{year} Order Schedule"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ["Date", "Day of Week", "Week Num", "Month", "Suppliers to Order", "Total Suppliers"]
    for col, header in enumerate(headers, 1):
        cell = ws_schedule.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Populate Dates
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    delta = timedelta(days=1)
    
    current_date = start_date
    row_idx = 2
    
    while current_date <= end_date:
        # Determine suppliers for this date
        day_of_week = current_date.weekday() # 0=Mon
        day_of_month = current_date.day
        
        suppliers_today = []
        
        # Weekly
        if weekly_suppliers[day_of_week]:
            suppliers_today.extend(weekly_suppliers[day_of_week])
            
        # Monthly
        if monthly_suppliers[day_of_month]:
            suppliers_today.extend(monthly_suppliers[day_of_month])
            
        # Clean names
        clean_sups = [s.replace('LIMITED', '').replace('LTD', '').strip().title() for s in suppliers_today]
        suppliers_str = ", ".join(clean_sups) if clean_sups else "-"
        # Filter Friendly Format: "Supplier A, Supplier B"
        
        # Write to Row
        ws_schedule.cell(row=row_idx, column=1, value=current_date)
        ws_schedule.cell(row=row_idx, column=2, value=current_date.strftime("%A"))
        ws_schedule.cell(row=row_idx, column=3, value=current_date.isocalendar()[1])
        ws_schedule.cell(row=row_idx, column=4, value=current_date.strftime("%B"))
        ws_schedule.cell(row=row_idx, column=5, value=suppliers_str)
        ws_schedule.cell(row=row_idx, column=6, value=len(clean_sups))
        
        # Styling for current row
        for col in range(1, 7):
            cell = ws_schedule.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col == 1:
                cell.number_format = 'DD-MMM-YYYY'
                
        current_date += delta
        row_idx += 1
        
    # Add AutoFilter
    ws_schedule.auto_filter.ref = f"A1:F{row_idx-1}"
    
    # Column Widths
    ws_schedule.column_dimensions['A'].width = 15
    ws_schedule.column_dimensions['B'].width = 15
    ws_schedule.column_dimensions['C'].width = 10
    ws_schedule.column_dimensions['D'].width = 15
    ws_schedule.column_dimensions['E'].width = 80 # Wide for suppliers
    ws_schedule.column_dimensions['F'].width = 15
    
    # --- SHEET 2: Daily & Legend ---
    ws_daily = wb.create_sheet("Daily Suppliers & Info")
    
    # Headers
    ws_daily.cell(row=1, column=1, value="Daily Orders (High Frequency)").font = Font(bold=True, size=14)
    ws_daily.cell(row=3, column=1, value="Supplier Name").font = header_font
    ws_daily.cell(row=3, column=1).fill = header_fill
    ws_daily.cell(row=3, column=2, value="Avg Lead Time (Days)").font = header_font
    ws_daily.cell(row=3, column=2).fill = header_fill
    
    for i, (name, lt) in enumerate(daily_suppliers):
        r = i + 4
        ws_daily.cell(row=r, column=1, value=name.title())
        ws_daily.cell(row=r, column=2, value=lt)
        
    ws_daily.column_dimensions['A'].width = 40
    ws_daily.column_dimensions['B'].width = 20
    
    # Irregular list
    if irregular_suppliers:
        base_row = len(daily_suppliers) + 6
        ws_daily.cell(row=base_row, column=1, value="Irregular / Bi-Weekly Suppliers").font = Font(bold=True, size=14)
        
        ws_daily.cell(row=base_row+2, column=1, value="Supplier Name").font = header_font
        ws_daily.cell(row=base_row+2, column=1).fill = header_fill
        ws_daily.cell(row=base_row+2, column=2, value="Category").font = header_font
        ws_daily.cell(row=base_row+2, column=2).fill = header_fill
        
        for i, (name, cat) in enumerate(irregular_suppliers):
            r = base_row + 3 + i
            ws_daily.cell(row=r, column=1, value=name.title())
            ws_daily.cell(row=r, column=2, value=cat)

    # Save
    wb.save(output_path)
    print(f"Excel Calendar generated at: {output_path}")

if __name__ == "__main__":
    out_file = os.path.join(os.getcwd(), f"Supplier_Order_Calendar_2026.xlsx")
    generate_excel_calendar(out_file)
