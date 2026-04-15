import openpyxl
import sys
import os

def extract_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Find the right worksheet
    ws = None
    for sheet in wb.worksheets:
        if sheet.title != "Order Summary":
            ws = sheet
            break
            
    if not ws: ws = wb.active
    
    data = {}
    
    filename = os.path.basename(file_path)
    print(f"File: {filename} (Sheet: {ws.title})")
            
    header_row_idx = None
    desc_col = None
    rec_col = None
    reason_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.upper()
                if 'DESCRIPTION' in val or 'PRODUCT' in val or 'ITEM' in val:
                    header_row_idx = cell.row
                    desc_col = cell.column
                if ('RECOMMENDED' in val and 'QTY' in val) or val == 'AI QUANTITY':
                    rec_col = cell.column
                if 'REASONING' in val or 'AI RATIONALE' in val:
                    reason_col = cell.column
        if desc_col: break
        
    if header_row_idx:
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row_idx, column=col).value
            if not v: continue
            val = str(v).upper()
            if 'RECOM' in val and 'QTY' in val: rec_col = col
            if 'REASONING' in val or 'AI REASONING' in val: reason_col = col
            if 'DESCRIPTION' in val or 'PRODUCT' in val: desc_col = col
            
    print(f"  Header Row: {header_row_idx}, Desc Col: {desc_col}, Rec Col: {rec_col}, Reason Col: {reason_col}")
    
    if desc_col:
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=desc_col).value
            if not name: continue
            qty = ws.cell(row=row_idx, column=rec_col).value if rec_col else 0
            reason = ws.cell(row=row_idx, column=reason_col).value if reason_col else ''
            data[str(name).strip()] = {'qty': qty, 'reason': reason}
    print(f"  Total Extracted Items: {len(data)}")
    return data

def compare_files(old_file, new_file):
    print(f"\n--- Comparing {os.path.basename(old_file).split('_')[1]} ---")
    try:
        old_data = extract_data(old_file)
        new_data = extract_data(new_file)
        
        diff_count = 0
        print("Differences:")
        for item, old_info in old_data.items():
            if item in new_data:
                new_info = new_data[item]
                o_qty = old_info['qty'] or 0
                n_qty = new_info['qty'] or 0
                try: o_qty = float(o_qty)
                except: o_qty = 0
                try: n_qty = float(n_qty)
                except: n_qty = 0
                
                if abs(o_qty - n_qty) > 0.01:
                    print(f"Item: {item}")
                    print(f"  OLD QTY: {o_qty} | REASON: {old_info['reason']}")
                    print(f"  NEW QTY: {n_qty} | REASON: {new_info['reason']}")
                    diff_count += 1
                    if diff_count > 10: 
                        print("  ... (showing first 10 differences only)")
                        break
        print(f"Total differing items shown: {diff_count}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    compare_files(r'C:\Users\iLink\Desktop\Projects\processed_unilever_1766597272.xlsx', r'C:\Users\iLink\Desktop\Projects\processed_unilever_1774024879.xlsx')
    compare_files(r'C:\Users\iLink\Desktop\Projects\processed_towfiq_1766597721.xlsx', r'C:\Users\iLink\Desktop\Projects\processed_towfiq_1774024847.xlsx')
