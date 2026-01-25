from openpyxl import load_workbook
import os

fpath = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data/prts_1.xlsx'
if os.path.exists(fpath):
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    for i, row in enumerate(rows):
        print(f"Row {i}: {' | '.join(str(v) or 'NONE' for v in row)}")
else:
    print(f"File not found: {fpath}")
