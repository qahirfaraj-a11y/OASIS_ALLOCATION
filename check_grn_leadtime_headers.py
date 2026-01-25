from openpyxl import load_workbook
import os

fpath = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data/grnds_10_10.5.xlsx'
if os.path.exists(fpath):
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    print(f"--- Checking {os.path.basename(fpath)} ---")
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    print(f"Headers: {header_row}")
    data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    print(f"Sample Row: {data_row}")
else:
    print(f"File not found: {fpath}")
