import asyncio
import os
import openpyxl
from app.logic.order_engine import OrderEngine

async def verify_flow():
    data_dir = os.path.join(os.getcwd(), "app", "data")
    input_file = "sample_picking_list.xlsx"
    output_file = "verified_recommendations.xlsx"
    
    if not os.path.exists(input_file):
        print("Input file missing. Run create_sample_file.py first.")
        return

    print("--- Starting End-to-End Verification ---")
    engine = OrderEngine(data_dir)
    
    # MOCK AI for local testing
    async def mock_ai(products, batch_num, total_batches):
        print(f"MOCK: Analyzing batch {batch_num} ({len(products)} products)")
        results = []
        for p in products:
            qty = 12 if "MILK" in p['product_name'] else 3
            results.append({
                "product_name": p['product_name'],
                "supplier_name": p['supplier_name'],
                "current_stock": p['current_stock'],
                "recommended_quantity": qty,
                "days_since_delivery": p['days_since_delivery'],
                "last_delivery_quantity": p['last_delivery_quantity'],
                "product_category": p['product_category'],
                "sales_velocity": p['sales_velocity'],
                "estimated_delivery_days": p['estimated_delivery_days'],
                "supplier_frequency": p['supplier_frequency'],
                "reorder_point": p['reorder_point'],
                "safety_stock_pct": p['safety_stock_pct'],
                "reasoning": "MOCKED REASONING"
            })
        return results
    
    engine.analyze_batch_ai = mock_ai

    try:
        results = await engine.run_intelligent_analysis(input_file, output_file)
        if results:
            print(f"SUCCESS: Analysis returned {len(results)} recommendations.")
            first_rec = results[0]
            required = ["product_name", "supplier_name", "current_stock", "recommended_quantity", 
                        "days_since_delivery", "last_delivery_quantity", "product_category", 
                        "sales_velocity", "estimated_delivery_days", "supplier_frequency", 
                        "reorder_point", "safety_stock_pct", "reasoning"]
            missing_fields = [f for f in required if f not in first_rec]
            if not missing_fields:
                print("SUCCESS: All 13 forecasting fields found in result dictionary.")
            else:
                print(f"FAILURE: Missing forecasting fields in dictionary: {missing_fields}")
        else:
            print("FAILURE: Analysis returned NO recommendations.")
    except Exception as e:
        print(f"Analysis failed: {e}")
        # Even if AI fails, let's see if the file was partially generated 
        # (Though current logic aborts)
        
    if os.path.exists(output_file):
        print(f"SUCCESS: Output file generated: {output_file}")
        
        wb = openpyxl.load_workbook(output_file)
        
        # Check Summary Sheet
        if "Order Summary" in wb.sheetnames:
            print("SUCCESS: 'Order Summary' sheet found.")
            ws_sum = wb["Order Summary"]
            print(f"Summary Row 3 (Cost): {ws_sum['A3'].value} = {ws_sum['B3'].value}")
        else:
            print("FAILURE: 'Order Summary' sheet MISSING.")
            
        # Check Data Sheet
        ws = wb["Picking List"] # Correct sheet name
        headers = [ws.cell(row=3, column=c).value for c in range(1, 40)] # Check more columns
        print(f"Headers found in output data sheet: {headers}")
        
        expected_cols = ["Recommended Qty", "Historical Avg", "Confidence", "Reasoning", "Est. Cost (KES)"]
        missing = [c for c in expected_cols if c not in headers]
        
        if not missing:
            print("SUCCESS: All 5 new columns found in data sheet.")
            # Check a row
            row4 = [ws.cell(row=4, column=c).value for c in range(1, 20)]
            print(f"Sample Row 4 (Data): {row4}")
        else:
            print(f"FAILURE: Missing columns: {missing}")
            
    else:
        print("FAILURE: Output file was NOT generated.")

if __name__ == "__main__":
    asyncio.run(verify_flow())
