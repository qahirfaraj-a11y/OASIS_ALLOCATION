from openpyxl import load_workbook
import os
import glob

data_dir = 'c:/Users/iLink/.gemini/antigravity/scratch/app/data'
files = glob.glob(os.path.join(data_dir, 'prts_*.xlsx'))

reasons = set()
for fpath in files[:3]: # Check first 3 files
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, max_row=100, values_only=True))
    # Reason is at index 8 (0-indexed) based on row 0: Org | Ven | Date | Doc | GRN | Barcode | Name | Status | Reason
    for row in rows:
        if len(row) > 8 and row[8]:
            reasons.add(str(row[8]).strip())

print(f"Discovered reasons: {reasons}")
