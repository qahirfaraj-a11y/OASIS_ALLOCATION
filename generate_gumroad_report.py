"""
OASIS Supplier Intelligence Report Generator v2
================================================
Generates a professional Excel report with DYNAMIC reliability scores.

Data Sources:
1. supplier_patterns_2025.json - Order frequency, base patterns
2. All_Suppliers_Fulfillment_Detail.xlsx - Actual PO->GRN fulfillment data
3. supplier_quality_scores_2025.json - Returns data (expiry, damaged, short supply)

Dynamic Reliability Formula:
  Score = (On-Time Rate * 0.35) + (Consistency * 0.25) + (Volume Confidence * 0.20) + (Quality * 0.20)
"""

import json
import os
import statistics
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Paths
DATA_DIR = os.path.dirname(__file__)
SUPPLIER_PATTERNS_PATH = os.path.join(DATA_DIR, "app", "data", "supplier_patterns_2025 (3).json")
FULFILLMENT_PATH = os.path.join(DATA_DIR, "All_Suppliers_Fulfillment_Detail.xlsx")
QUALITY_PATH = os.path.join(DATA_DIR, "app", "data", "supplier_quality_scores_2025 (1).json")
OUTPUT_PATH = os.path.join(DATA_DIR, "Supplier_Intelligence_Report_2025_v3.xlsx")


def load_json(path):
    """Load JSON file."""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def analyze_fulfillment_data():
    """
    Analyze fulfillment Excel for dynamic metrics:
    - Average fulfillment days
    - Standard deviation (consistency)
    - On-time rate (% within expected lead time)
    - Total PO count
    
    Uses Sheet1 which contains full data (107k+ rows).
    """
    wb = load_workbook(FULFILLMENT_PATH, data_only=True)
    
    # Use Sheet1 which has the full data (107k rows)
    ws = wb['Sheet1']
    
    # Sheet1 columns: Vendor Name(1), PO No(2), PO Date(3), GRN No(4), GRN Date(5), 
    #                 Fulfillment Days(6), Item Name(7), GRN Qty(8), Net Amt(9)
    
    supplier_stats = defaultdict(lambda: {
        'po_set': set(),
        'grn_set': set(),
        'fulfillment_days': [],
        'total_value': 0.0,
        'line_items': 0
    })
    
    print(f"  Reading {ws.max_row:,} rows from Sheet1...")
    
    for row in range(2, ws.max_row + 1):
        vendor_raw = ws.cell(row, 1).value
        if not vendor_raw:
            continue
        
        # Normalize vendor name (remove code prefix like "SA0029 - ")
        vendor = str(vendor_raw).split(" - ", 1)[-1].strip().upper() if " - " in str(vendor_raw) else str(vendor_raw).strip().upper()
        
        po_no = ws.cell(row, 2).value
        grn_no = ws.cell(row, 4).value
        f_days = ws.cell(row, 6).value  # Column 6 in Sheet1
        net_amt = ws.cell(row, 9).value or 0  # Column 9 in Sheet1
        
        stats = supplier_stats[vendor]
        if po_no:
            stats['po_set'].add(po_no)
        if grn_no:
            stats['grn_set'].add(grn_no)
        if f_days and isinstance(f_days, (int, float)):
            stats['fulfillment_days'].append(int(f_days))
        stats['total_value'] += float(net_amt) if net_amt else 0
        stats['line_items'] += 1
    
    # Convert to metrics
    result = {}
    for vendor, stats in supplier_stats.items():
        days = stats['fulfillment_days']
        if not days:
            continue
        
        avg_days = statistics.mean(days)
        std_days = statistics.stdev(days) if len(days) > 1 else 0
        min_days = min(days)
        max_days = max(days)
        
        result[vendor] = {
            'total_pos': len(stats['po_set']),
            'total_grns': len(stats['grn_set']),
            'total_value': stats['total_value'],
            'avg_fulfillment_days': round(avg_days, 1),
            'std_fulfillment_days': round(std_days, 2),
            'min_days': min_days,
            'max_days': max_days,
            'line_items': stats['line_items']
        }
    
    return result


def calculate_dynamic_reliability(patterns: dict, fulfillment: dict, quality: dict) -> dict:
    """
    Calculate dynamic reliability score for each supplier.
    
    Formula (weighted average of 4 components):
    1. On-Time Rate (35%): Based on fulfillment days vs expected lead time
    2. Consistency (25%): Based on std deviation of fulfillment days
    3. Volume Confidence (20%): More POs = higher confidence
    4. Quality Score (20%): From returns data
    """
    results = {}
    
    for name, pat in patterns.items():
        # Base values from patterns (fallback)
        patterns_lead_time = pat.get('avg_gap_days', pat.get('estimated_delivery_days', 7))
        total_orders = pat.get('total_orders_2025', 0)
        
        # Check fulfillment data - PRIMARY SOURCE for lead times
        ful = fulfillment.get(name, {})
        if not ful:
            # Try partial match (fuzzy)
            for ful_name, ful_data in fulfillment.items():
                if name in ful_name or ful_name in name:
                    ful = ful_data
                    break
        
        # Check quality data
        qual = quality.get(name, {})
        if not qual:
            for qual_name, qual_data in quality.items():
                if name in qual_name or qual_name in name:
                    qual = qual_data
                    break
        
        # Determine lead time source (prioritize fulfillment data)
        has_fulfillment_data = ful and ful.get('avg_fulfillment_days') is not None
        if has_fulfillment_data:
            actual_lead_time = ful['avg_fulfillment_days']
            lead_time_std = ful.get('std_fulfillment_days', 0)
            data_source = 'Fulfillment'
            po_count = ful.get('total_pos', total_orders)
        else:
            actual_lead_time = patterns_lead_time
            lead_time_std = 0  # Unknown
            data_source = 'Estimated'
            po_count = total_orders
        
        # Component 1: On-Time Rate (35%)
        # Compare actual lead time to category benchmark
        if has_fulfillment_data:
            # Actual data: score based on how fast they deliver
            if actual_lead_time <= 3:
                on_time_rate = 1.0  # Excellent
            elif actual_lead_time <= 7:
                on_time_rate = 0.95  # Good
            elif actual_lead_time <= 14:
                on_time_rate = 0.85  # Average
            else:
                on_time_rate = max(0.6, 0.80 - ((actual_lead_time - 14) / 50))  # Slow
        else:
            # No actual data - use neutral estimate
            on_time_rate = 0.80
        
        # Component 2: Consistency (25%)
        # Low std = high consistency
        if has_fulfillment_data and lead_time_std is not None:
            std = lead_time_std
            avg = actual_lead_time or 1
            cv = std / avg if avg > 0 else 0.5  # Coefficient of variation
            consistency = max(0.5, 1 - cv)  # Lower CV = higher score
        else:
            consistency = 0.75  # Default (lower for no data)
        
        # ============================================
        # VOLUME MULTIPLIER (applies to quality and final score)
        # ============================================
        # High-volume suppliers get credit for scale and managing complexity
        frequency = pat.get('order_frequency', 'monthly')
        
        # Volume tier based on orders per year
        if po_count >= 200:
            volume_multiplier = 1.25  # Very high volume (daily suppliers)
        elif po_count >= 100:
            volume_multiplier = 1.15  # High volume
        elif po_count >= 50:
            volume_multiplier = 1.10  # Medium-high
        elif po_count >= 20:
            volume_multiplier = 1.0   # Normal
        else:
            volume_multiplier = 0.95  # Low volume (less data confidence)
        
        # Frequency bonus (daily is harder to manage consistently)
        if frequency == 'daily':
            frequency_bonus = 1.10
        elif frequency in ['every_2_3_days', 'every 2 3 days']:
            frequency_bonus = 1.05
        elif frequency == 'weekly':
            frequency_bonus = 1.0
        else:
            frequency_bonus = 0.98
        
        # Combined volume multiplier
        combined_volume_mult = min(1.4, volume_multiplier * frequency_bonus)
        
        # Component 3: Volume Confidence (20%)
        # More POs = more confidence in the data
        if po_count >= 50:
            volume_confidence = 1.0
        elif po_count >= 20:
            volume_confidence = 0.9
        elif po_count >= 10:
            volume_confidence = 0.8
        elif po_count >= 5:
            volume_confidence = 0.7
        else:
            volume_confidence = 0.6
        
        # Component 4: Quality Score (20%)
        # RATE-BASED with VOLUME MULTIPLIER
        # Component 4: Quality Score (20%)
        # VALUE-BASED: Use return value % vs total purchase value (most accurate commercial metric)
        if qual:
            total_return_value = qual.get('total_value_returned', 0)
            total_purch_value = ful.get('total_value', 0)
            
            # Fallback to estimated value if missing (avg order value * count)
            if total_purch_value == 0:
                avg_val = pat.get('avg_order_value_kes', 5000)
                total_purch_value = avg_val * po_count
            
            # Calculate Return Rate by Value (%)
            if total_purch_value > 0:
                return_rate_pct = (total_return_value / total_purch_value) * 100
                
                # Special handling for massive return rates (likely crate/bottle deposits like Coca Cola)
                if return_rate_pct > 25 and total_purch_value > 1000000:
                   # Cap the effective rate for scoring purposes - assume it's deposits
                   effective_rate = 15 + (return_rate_pct - 25) * 0.1
                else:
                   effective_rate = return_rate_pct
            else:
                effective_rate = 0
            
            # Score based on commercial standards:
            # < 1% = Excellent (100%)
            # 1-3% = Good (90-100%)
            # 3-5% = Fair (80-90%) - DPL/Bio Food fall here
            # 5-8% = Acceptable (70-80%) - Brookside falls here
            # > 8% = Poor (< 70%)
            
            if effective_rate <= 1.0:
                 base_quality = 1.0
            elif effective_rate <= 3.0:
                 base_quality = 1.0 - ((effective_rate - 1.0) / 20)  # 0.90 - 1.00
            elif effective_rate <= 5.0:
                 base_quality = 0.90 - ((effective_rate - 3.0) / 20) # 0.80 - 0.90
            elif effective_rate <= 8.0:
                 base_quality = 0.80 - ((effective_rate - 5.0) / 30) # 0.70 - 0.80
            elif effective_rate <= 15.0:
                 base_quality = 0.70 - ((effective_rate - 8.0) / 35) # 0.50 - 0.70
            else:
                 base_quality = 0.50
            
            # Apply multipliers
            quality_score = min(1.0, base_quality * combined_volume_mult)
            
            # Floor calculated score at 50%
            quality_score = max(0.50, quality_score)
            
        else:
            quality_score = 0.90  # Default if no returns data
        
        # Weighted average (base reliability)
        base_reliability = (
            on_time_rate * 0.35 +
            consistency * 0.25 +
            volume_confidence * 0.20 +
            quality_score * 0.20
        )
        
        # Apply volume multiplier to final reliability (capped boost)
        reliability = min(0.98, base_reliability * (1 + (combined_volume_mult - 1) * 0.3))
        
        # Determine risk level
        if reliability >= 0.85:
            risk = "Low"
        elif reliability >= 0.70:
            risk = "Medium"
        elif reliability >= 0.50:
            risk = "High"
        else:
            risk = "Critical"
        
        results[name] = {
            'reliability_score': round(reliability * 100, 1),
            'on_time_rate': round(on_time_rate * 100, 1),
            'consistency_score': round(consistency * 100, 1),
            'volume_confidence': round(volume_confidence * 100, 1),
            'quality_score': round(quality_score * 100, 1),
            'risk_level': risk,
            'total_pos': po_count,
            'avg_lead_time': ful.get('avg_fulfillment_days', pat.get('avg_gap_days', 7)),
            'std_lead_time': ful.get('std_fulfillment_days', 0),
            'total_value': ful.get('total_value', 0),
            'frequency': pat.get('order_frequency', 'unknown').replace('_', ' ').title(),
            # Quality details
            'expiry_returns': qual.get('expiry_returns', 0),
            'damaged_returns': qual.get('damaged_returns', 0),
            'short_supply': qual.get('short_supply_returns', 0)
        }
    
    return results


def generate_report():
    """Generate the Gumroad-ready Excel report with dynamic scores."""
    print("Loading data sources...")
    patterns = load_json(SUPPLIER_PATTERNS_PATH)
    quality = load_json(QUALITY_PATH)
    
    print("Analyzing fulfillment data...")
    fulfillment = analyze_fulfillment_data()
    
    print(f"Data: {len(patterns)} suppliers in patterns, {len(fulfillment)} with fulfillment data, {len(quality)} with quality data")
    
    print("Calculating dynamic reliability scores...")
    scored = calculate_dynamic_reliability(patterns, fulfillment, quality)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=18, color="1E3A5F")
    subtitle_font = Font(bold=True, size=14, color="4A90A4")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    risk_colors = {
        "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "High": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Critical": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    }
    
    # ===== SHEET 1: EXECUTIVE SUMMARY =====
    ws = wb.active
    ws.title = "Executive Summary"
    
    ws['A1'] = "OASIS Supplier Intelligence Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')} | Data Period: 2025"
    ws['A2'].font = Font(italic=True, color="666666")
    
    ws['A4'] = "SCORING METHODOLOGY"
    ws['A4'].font = subtitle_font
    
    methodology = [
        ("On-Time Delivery", "35%", "Actual fulfillment days vs expected lead time"),
        ("Consistency", "25%", "Standard deviation of delivery times (lower = better)"),
        ("Volume Confidence", "20%", "Number of POs (more data = higher confidence)"),
        ("Quality Score", "20%", "Returns rate (expiry, damaged, short supply)")
    ]
    
    ws['A5'] = "Component"
    ws['B5'] = "Weight"
    ws['C5'] = "Description"
    for col in ['A', 'B', 'C']:
        ws[f'{col}5'].font = header_font
        ws[f'{col}5'].fill = header_fill
    
    for i, (comp, weight, desc) in enumerate(methodology, 6):
        ws[f'A{i}'] = comp
        ws[f'B{i}'] = weight
        ws[f'C{i}'] = desc
    
    # Summary stats
    ws['A11'] = "SUMMARY STATISTICS"
    ws['A11'].font = subtitle_font
    
    total_suppliers = len(scored)
    low_risk = sum(1 for s in scored.values() if s['risk_level'] == 'Low')
    high_risk = sum(1 for s in scored.values() if s['risk_level'] in ['High', 'Critical'])
    avg_reliability = sum(s['reliability_score'] for s in scored.values()) / total_suppliers if total_suppliers else 0
    
    stats = [
        ("Total Suppliers Analyzed", total_suppliers),
        ("Average Reliability Score", f"{avg_reliability:.1f}%"),
        ("Low Risk Suppliers", low_risk),
        ("High/Critical Risk Suppliers", high_risk)
    ]
    
    for i, (label, value) in enumerate(stats, 12):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    
    # ===== SHEET 2: FULL RANKINGS =====
    ws_rank = wb.create_sheet("Supplier Rankings")
    
    headers = [
        "Rank", "Supplier Name", "Reliability", "On-Time", "Consistency",
        "Volume Conf.", "Quality", "Risk Level", "Avg Lead Time", "Total POs", "Frequency"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_rank.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Sort by reliability score
    sorted_suppliers = sorted(scored.items(), key=lambda x: -x[1]['reliability_score'])
    
    for i, (name, data) in enumerate(sorted_suppliers, 1):
        row = i + 1
        ws_rank.cell(row=row, column=1, value=i).border = thin_border
        ws_rank.cell(row=row, column=2, value=name[:45]).border = thin_border
        ws_rank.cell(row=row, column=3, value=f"{data['reliability_score']:.1f}%").border = thin_border
        ws_rank.cell(row=row, column=4, value=f"{data['on_time_rate']:.0f}%").border = thin_border
        ws_rank.cell(row=row, column=5, value=f"{data['consistency_score']:.0f}%").border = thin_border
        ws_rank.cell(row=row, column=6, value=f"{data['volume_confidence']:.0f}%").border = thin_border
        ws_rank.cell(row=row, column=7, value=f"{data['quality_score']:.0f}%").border = thin_border
        
        risk_cell = ws_rank.cell(row=row, column=8, value=data['risk_level'])
        risk_cell.fill = risk_colors.get(data['risk_level'], PatternFill())
        risk_cell.border = thin_border
        risk_cell.alignment = Alignment(horizontal='center')
        
        ws_rank.cell(row=row, column=9, value=f"{data['avg_lead_time']:.1f}d").border = thin_border
        ws_rank.cell(row=row, column=10, value=data['total_pos']).border = thin_border
        ws_rank.cell(row=row, column=11, value=data['frequency']).border = thin_border
    
    # Column widths
    widths = [6, 45, 12, 10, 12, 12, 10, 12, 12, 10, 14]
    for i, w in enumerate(widths, 1):
        ws_rank.column_dimensions[chr(64+i)].width = w
    
    # ===== SHEET 3: QUALITY ISSUES =====
    ws_qual = wb.create_sheet("Quality Analysis")
    
    ws_qual['A1'] = "Suppliers with Quality Issues"
    ws_qual['A1'].font = title_font
    
    headers = ["Supplier", "Quality Score", "Expiry Returns", "Damaged", "Short Supply", "Risk Level"]
    for col, header in enumerate(headers, 1):
        cell = ws_qual.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Filter to suppliers with quality issues
    quality_issues = [(n, d) for n, d in sorted_suppliers if d['quality_score'] < 90]
    quality_issues.sort(key=lambda x: x[1]['quality_score'])
    
    for i, (name, data) in enumerate(quality_issues[:50], 4):
        ws_qual.cell(row=i, column=1, value=name[:40])
        ws_qual.cell(row=i, column=2, value=f"{data['quality_score']:.0f}%")
        ws_qual.cell(row=i, column=3, value=data['expiry_returns'])
        ws_qual.cell(row=i, column=4, value=data['damaged_returns'])
        ws_qual.cell(row=i, column=5, value=data['short_supply'])
        
        risk_cell = ws_qual.cell(row=i, column=6, value=data['risk_level'])
        risk_cell.fill = risk_colors.get(data['risk_level'], PatternFill())
    
    ws_qual.column_dimensions['A'].width = 40
    
    # ===== SHEET 4: TOP 20 =====
    ws_top = wb.create_sheet("Top 20 Performers")
    
    ws_top['A1'] = "Top 20 Most Reliable Suppliers"
    ws_top['A1'].font = title_font
    
    headers = ["Rank", "Supplier", "Reliability", "Lead Time", "Consistency", "Recommendation"]
    for col, header in enumerate(headers, 1):
        cell = ws_top.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for i, (name, data) in enumerate(sorted_suppliers[:20], 4):
        ws_top.cell(row=i, column=1, value=i-3)
        ws_top.cell(row=i, column=2, value=name[:40])
        ws_top.cell(row=i, column=3, value=f"{data['reliability_score']:.1f}%")
        ws_top.cell(row=i, column=4, value=f"{data['avg_lead_time']:.1f} days")
        ws_top.cell(row=i, column=5, value=f"{data['consistency_score']:.0f}%")
        
        # Recommendation
        if data['avg_lead_time'] <= 2:
            rec = "Ideal for daily orders"
        elif data['avg_lead_time'] <= 7:
            rec = "Reliable for weekly orders"
        else:
            rec = "Plan with buffer stock"
        ws_top.cell(row=i, column=6, value=rec)
    
    ws_top.column_dimensions['A'].width = 6
    ws_top.column_dimensions['B'].width = 40
    ws_top.column_dimensions['F'].width = 28
    
    # Save
    wb.save(OUTPUT_PATH)
    print(f"\n[SUCCESS] Report saved to: {OUTPUT_PATH}")
    print(f"   - {len(scored)} suppliers with DYNAMIC reliability scores")
    print(f"   - {low_risk} Low Risk | {high_risk} High/Critical Risk")
    print(f"   - Average reliability: {avg_reliability:.1f}%")
    
    return OUTPUT_PATH


if __name__ == "__main__":
    generate_report()
